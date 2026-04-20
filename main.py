import csv
import getpass
import calendar
import json
import os
import re
import shutil
import subprocess
import sys
import threading
import uuid
import webbrowser
from collections import defaultdict
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Callable

import tkinter as tk
from tkinter import colorchooser, filedialog, simpledialog, ttk

import jdatetime
from openpyxl import Workbook, load_workbook
from PIL import Image, ImageDraw
import pystray
import ttkbootstrap as tb
from plyer import notification
from ttkbootstrap.constants import BOTH, END, LEFT, RIGHT, X, Y
from ttkbootstrap.dialogs import Messagebox
from winotify import Notification, Registry, audio

try:
    import winreg
except ImportError:  # pragma: no cover
    winreg = None

try:
    import winsound
except ImportError:  # pragma: no cover
    winsound = None


APP_TITLE = "Offline CRM"
APP_VERSION = "2"
APP_NOTIFY_ID = "CRM_Sales_App"
APP_ICON_CANDIDATES = [
    Path(__file__).resolve().parent / "Icon.png",
    Path.home() / "Desktop" / "Icon.png",
]
CUSTOMER_FIELDS = ["id", "name", "mobile", "birthdate", "income_level", "birthday_notified_for", "birthday_seen_for", "created_at", "updated_at"]
DEAL_FIELDS = [
    "id",
    "customer_id",
    "customer_name",
    "title",
    "deal_type",
    "product",
    "category",
    "pipeline",
    "status",
    "sale_price",
    "operator_commission",
    "sales_expert",
    "notes",
    "reminder_at",
    "last_notified_at",
    "notification_seen_at",
    "created_at",
    "updated_at",
]
CATEGORY_FIELDS = ["id", "title", "color"]
PIPELINE_FIELDS = ["id", "title"]
DEAL_TYPE_FIELDS = ["id", "title"]
PRODUCT_FIELDS = ["id", "deal_type", "title"]
PROFILE_FIELDS = ["id", "title", "username", "password", "created_at"]
DEFAULT_DEAL_TYPES = ["دوره", "تجهیزات", "تعمیرات"]
DEAL_STATUSES = ["در دست بررسی", "موفق", "ناموفق"]
DEFAULT_CATEGORIES = ["سرنخ جدید", "تماس اولیه", "پیگیری گرم", "جلسه فروش", "پیشنهاد قیمت", "مذاکره", "بسته شده"]
DEFAULT_PIPELINES = ["ورودی وب", "ورودی تلفنی", "ارجاعی", "همکاری سازمانی"]
INCOME_LEVELS = ["ضعیف", "متوسط", "خوب", "عالی"]
DEFAULT_CATEGORY_COLORS = ["#e76f51", "#f4a261", "#e9c46a", "#2a9d8f", "#457b9d", "#8d99ae", "#9c6644"]
STATUS_META = {
    "در دست بررسی": {"emoji": "ðŸŸ¨", "color": "#f4a261"},
    "موفق": {"emoji": "🟩", "color": "#2a9d8f"},
    "ناموفق": {"emoji": "🟥", "color": "#e76f51"},
}
DATASET_CONFIG = {
    "مشتریان": ("customers.csv", CUSTOMER_FIELDS),
    "معاملات": ("deals.csv", DEAL_FIELDS),
    "کاریزها": ("pipelines.csv", PIPELINE_FIELDS),
    "دسته‌بندی‌های معامله": ("categories.csv", CATEGORY_FIELDS),
    "نوع معامله": ("deal_types.csv", DEAL_TYPE_FIELDS),
    "محصولات نوع معامله": ("products.csv", PRODUCT_FIELDS),
}
PERSIAN_MONTHS = [
    "فروردین",
    "اردیبهشت",
    "خرداد",
    "تیر",
    "مرداد",
    "شهریور",
    "مهر",
    "آبان",
    "آذر",
    "دی",
    "بهمن",
    "اسفند",
]
WINDOWS_RUN_KEY = r"Software\Microsoft\Windows\CurrentVersion\Run"
WINDOWS_RUN_VALUE = "OfflineCRM"
APP_SETTINGS_FILENAME = "app_settings.json"


def default_local_appdata_root() -> Path:
    raw = os.environ.get("LOCALAPPDATA")
    if raw:
        return Path(raw)
    return Path.home() / "AppData" / "Local"


def runtime_base_dir() -> Path:
    if getattr(sys, "frozen", False):
        return Path(sys.executable).resolve().parent
    return Path(__file__).resolve().parent


def persistent_data_root() -> Path:
    if getattr(sys, "frozen", False):
        return default_local_appdata_root() / "OfflineCRM" / "data"
    return runtime_base_dir() / "data"


def legacy_data_root_candidates() -> list[Path]:
    candidates: list[Path] = []
    for path in [runtime_base_dir() / "data", Path(__file__).resolve().parent / "data"]:
        if path not in candidates:
            candidates.append(path)
    return candidates


DATA_ROOT = persistent_data_root()


def migrate_legacy_data_if_needed() -> None:
    target = DATA_ROOT
    target.parent.mkdir(parents=True, exist_ok=True)
    if target.exists() and any(target.iterdir()):
        return
    for legacy_root in legacy_data_root_candidates():
        if legacy_root == target or not legacy_root.exists() or not any(legacy_root.iterdir()):
            continue
        shutil.copytree(legacy_root, target, dirs_exist_ok=True)
        return
def repair_text(value):
    if not isinstance(value, str) or not value:
        return value
    suspicious_markers = "ÃØÙÚÛŒÂâð"
    if not any(ch in value for ch in suspicious_markers):
        return value

    def suspicious_count(text: str) -> int:
        return sum(text.count(ch) for ch in suspicious_markers)

    best = value
    for _ in range(3):
        try:
            candidate = best.encode("cp1252").decode("utf-8")
        except Exception:
            break
        if suspicious_count(candidate) < suspicious_count(best) or any("\u0600" <= ch <= "\u06FF" for ch in candidate):
            best = candidate
        else:
            break
    if best != value:
        return best

    pattern = re.compile(r"[ÃØÙÚÛŒÂâð][^{}\n\r]*")

    def fix_segment(segment: str) -> str:
        best = segment
        for _ in range(3):
            try:
                candidate = best.encode("cp1252").decode("utf-8")
            except Exception:
                break
            if suspicious_count(candidate) < suspicious_count(best) or any("\u0600" <= ch <= "\u06FF" for ch in candidate):
                best = candidate
            else:
                break
        return best

    repaired = pattern.sub(lambda match: fix_segment(match.group(0)), value)
    return repaired


def repair_structure(value):
    if isinstance(value, str):
        return repair_text(value)
    if isinstance(value, list):
        return [repair_structure(item) for item in value]
    if isinstance(value, tuple):
        return tuple(repair_structure(item) for item in value)
    if isinstance(value, dict):
        return {repair_structure(key): repair_structure(item) for key, item in value.items()}
    return value


DEFAULT_DEAL_TYPES = repair_structure(DEFAULT_DEAL_TYPES)
DEAL_STATUSES = repair_structure(DEAL_STATUSES)
DEFAULT_CATEGORIES = repair_structure(DEFAULT_CATEGORIES)
DEFAULT_PIPELINES = repair_structure(DEFAULT_PIPELINES)
INCOME_LEVELS = repair_structure(INCOME_LEVELS)
STATUS_META = repair_structure(STATUS_META)
DATASET_CONFIG = repair_structure(DATASET_CONFIG)
PERSIAN_MONTHS = repair_structure(PERSIAN_MONTHS)


def repair_widget_kwargs(kwargs: dict) -> dict:
    repaired = dict(kwargs)
    if "text" in repaired:
        repaired["text"] = repair_text(repaired["text"])
    if "title" in repaired:
        repaired["title"] = repair_text(repaired["title"])
    if "values" in repaired:
        repaired["values"] = repair_structure(repaired["values"])
    return repaired


def patch_widget_text_repair():
    widget_classes = [tb.Button, tb.Checkbutton, tb.Label, tb.LabelFrame, ttk.Label, ttk.Labelframe]
    for widget_class in widget_classes:
        original_init = widget_class.__init__

        def patched_init(self, *args, __orig=original_init, **kwargs):
            return __orig(self, *args, **repair_widget_kwargs(kwargs))

        widget_class.__init__ = patched_init

    original_heading = ttk.Treeview.heading

    def patched_heading(self, column, option=None, **kwargs):
        if "text" in kwargs:
            kwargs["text"] = repair_text(kwargs["text"])
        return original_heading(self, column, option=option, **kwargs)

    ttk.Treeview.heading = patched_heading

    original_notebook_add = ttk.Notebook.add

    def patched_notebook_add(self, child, **kwargs):
        if "text" in kwargs:
            kwargs["text"] = repair_text(kwargs["text"])
        return original_notebook_add(self, child, **kwargs)

    ttk.Notebook.add = patched_notebook_add

    original_combobox_init = ttk.Combobox.__init__

    def patched_combobox_init(self, master=None, **kwargs):
        kwargs = repair_widget_kwargs(kwargs)
        return original_combobox_init(self, master=master, **kwargs)

    ttk.Combobox.__init__ = patched_combobox_init

    original_title = tk.Wm.title

    def patched_title(self, string=None):
        if string is None:
            return original_title(self)
        return original_title(self, repair_text(string))

    tk.Wm.title = patched_title

    for method_name in ["show_error", "show_info", "show_warning", "okcancel", "yesno"]:
        original_method = getattr(Messagebox, method_name)

        def patched_messagebox(*args, __orig=original_method, **kwargs):
            fixed_args = list(args)
            if fixed_args:
                fixed_args[0] = repair_text(fixed_args[0])
            if len(fixed_args) > 1:
                fixed_args[1] = repair_text(fixed_args[1])
            return __orig(*fixed_args, **kwargs)

        setattr(Messagebox, method_name, patched_messagebox)


patch_widget_text_repair()


def now_text() -> str:
    return now_jalali_text()


def to_float(value: str) -> float:
    try:
        return float(str(value).replace(",", "").strip() or 0)
    except ValueError:
        return 0.0


def safe_slug(value: str) -> str:
    cleaned = "".join(ch if ch.isalnum() else "_" for ch in value.strip())
    return cleaned.strip("_") or "profile"


def rtl_label(parent, text: str, **kwargs):
    return tb.Label(parent, text=repair_text(text), anchor="e", justify="right", **kwargs)


def rtl_text(text: str) -> str:
    return f"\u202B{repair_text(text)}\u202C"


def get_app_icon_path() -> Path | None:
    for path in APP_ICON_CANDIDATES:
        if path.exists():
            return path
    return None


def now_jalali_text() -> str:
    return jdatetime.datetime.now().strftime("%Y/%m/%d %H:%M:%S")


def app_settings_path_for_profile(profile: "ProfileContext") -> Path:
    return profile.root.parent / APP_SETTINGS_FILENAME


def load_app_settings(settings_path: Path) -> dict:
    if not settings_path.exists():
        return {"launch_on_startup": True}
    try:
        return json.loads(settings_path.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError):
        return {"launch_on_startup": True}


def save_app_settings(settings_path: Path, settings: dict) -> None:
    settings_path.write_text(json.dumps(settings, ensure_ascii=False, indent=2), encoding="utf-8")


def startup_command() -> str:
    if getattr(sys, "frozen", False):
        return f'"{sys.executable}"'
    return f'"{sys.executable}" "{Path(__file__).resolve()}"'


def set_windows_startup(enabled: bool) -> bool:
    if winreg is None:
        return False
    try:
        with winreg.OpenKey(winreg.HKEY_CURRENT_USER, WINDOWS_RUN_KEY, 0, winreg.KEY_SET_VALUE) as key:
            if enabled:
                winreg.SetValueEx(key, WINDOWS_RUN_VALUE, 0, winreg.REG_SZ, startup_command())
            else:
                try:
                    winreg.DeleteValue(key, WINDOWS_RUN_VALUE)
                except FileNotFoundError:
                    pass
        return True
    except OSError:
        return False


def shorten_text(value: str, limit: int = 70) -> str:
    cleaned = " ".join((value or "").split())
    if len(cleaned) <= limit:
        return cleaned
    return f"{cleaned[: limit - 1].rstrip()}..."


RICH_NOTE_PREFIX = "__RICH_TEXT__:"


def decode_rich_note(value: str) -> dict | None:
    if not isinstance(value, str) or not value.startswith(RICH_NOTE_PREFIX):
        return None
    try:
        return json.loads(value[len(RICH_NOTE_PREFIX) :])
    except json.JSONDecodeError:
        return None


def note_plain_text(value: str) -> str:
    payload = decode_rich_note(value or "")
    if payload:
        return payload.get("text", "")
    return value or ""


def note_preview_text(value: str, limit: int = 90) -> str:
    lines = [line.strip() for line in note_plain_text(value).splitlines() if line.strip()]
    if not lines:
        return ""
    chosen = ""
    for line in reversed(lines):
        if re.fullmatch(r"\[\d{4}/\d{2}/\d{2} \d{2}:\d{2}:\d{2}\]", line):
            continue
        chosen = line
        break
    chosen = chosen or lines[-1]
    chosen = re.sub(r"^[>\-\*\d\.\[\]\(\)\s#]+", "", chosen).strip()
    return shorten_text(chosen, limit=limit)


def parse_notes_history(value: str) -> list[dict[str, str]]:
    text = (value or "").strip()
    if not text:
        return []
    entries: list[dict[str, str]] = []
    current_time = ""
    current_lines: list[str] = []
    timestamp_pattern = re.compile(r"^\[(\d{4}/\d{2}/\d{2} \d{2}:\d{2}:\d{2})\]$")
    for raw_line in text.splitlines():
        line = raw_line.rstrip()
        match = timestamp_pattern.match(line.strip())
        if match:
            if current_time or any(part.strip() for part in current_lines):
                entries.append(
                    {
                        "time": current_time or "بدون زمان",
                        "text": "\n".join(current_lines).strip(),
                    }
                )
            current_time = match.group(1)
            current_lines = []
            continue
        current_lines.append(line)
    if current_time or any(part.strip() for part in current_lines):
        entries.append(
            {
                "time": current_time or "بدون زمان",
                "text": "\n".join(current_lines).strip(),
            }
        )
    cleaned_entries = []
    for entry in entries:
        preview = note_preview_text(entry["text"], limit=120)
        cleaned_entries.append(
            {
                "time": entry["time"],
                "text": entry["text"] or preview or "-",
                "preview": preview or "-",
            }
        )
    return cleaned_entries


def serialize_notes_history(entries: list[dict[str, str]]) -> str:
    chunks = []
    for entry in entries:
        text = (entry.get("text") or "").strip()
        if not text:
            continue
        timestamp = (entry.get("time") or now_text()).strip()
        chunks.append(f"[{timestamp}]\n{text}")
    return "\n\n".join(chunks).strip()


def normalize_keyboard_text(value: str) -> str:
    mapping = str.maketrans(
        {
            "ي": "ی",
            "ك": "ک",
            "ة": "ه",
            "ۀ": "ه",
            "ؤ": "و",
            "إ": "ا",
            "أ": "ا",
            "٠": "0",
            "١": "1",
            "٢": "2",
            "٣": "3",
            "٤": "4",
            "٥": "5",
            "٦": "6",
            "٧": "7",
            "٨": "8",
            "٩": "9",
            "۰": "0",
            "۱": "1",
            "۲": "2",
            "۳": "3",
            "۴": "4",
            "۵": "5",
            "۶": "6",
            "۷": "7",
            "۸": "8",
            "۹": "9",
        }
    )
    return (value or "").translate(mapping)


def normalize_persian_editor_text(value: str) -> str:
    mapping = str.maketrans(
        {
            "ي": "ی",
            "ك": "ک",
            "ة": "ه",
            "ۀ": "ه",
            "ؤ": "و",
            "إ": "ا",
            "أ": "ا",
            "ئ": "ی",
            "‌ ": " ",
        }
    )
    return (value or "").translate(mapping)


def matches_search(query: str, *values: str) -> bool:
    normalized_query = normalize_keyboard_text(query).strip()
    if not normalized_query:
        return True
    return any(normalized_query in normalize_keyboard_text(value or "") for value in values)


def widget_supports_clipboard(widget) -> bool:
    return widget.winfo_class() in {"Entry", "TEntry", "Text", "Combobox", "TCombobox", "Spinbox", "TSpinbox"}


def show_edit_context_menu(event, menu: tk.Menu):
    widget = event.widget
    if not widget_supports_clipboard(widget):
        return
    try:
        widget.focus_force()
    except Exception:
        pass
    menu.target_widget = widget
    state = "normal"
    try:
        selection = widget.selection_get()
        has_selection = bool(selection)
    except Exception:
        has_selection = False
    menu.entryconfigure("برش", state=state if has_selection else "disabled")
    menu.entryconfigure("کپی", state=state if has_selection else "disabled")
    menu.entryconfigure("چسباندن", state=state)
    menu.entryconfigure("انتخاب همه", state=state)
    menu.tk_popup(event.x_root, event.y_root)


def run_context_action(menu: tk.Menu, virtual_event: str):
    widget = getattr(menu, "target_widget", None)
    if widget is None:
        return
    if virtual_event == "<<SelectAll>>":
        try:
            if widget.winfo_class() in {"Text"}:
                widget.tag_add("sel", "1.0", "end-1c")
                widget.mark_set("insert", "end-1c")
            else:
                widget.selection_range(0, END)
                widget.icursor(END)
            return
        except Exception:
            return
    try:
        widget.event_generate(virtual_event)
    except Exception:
        pass


def bind_edit_shortcuts(window):
    if getattr(window, "_edit_shortcuts_bound", False):
        return
    window._edit_shortcuts_bound = True
    menu = tk.Menu(window, tearoff=0)
    menu.target_widget = None
    menu.add_command(label="برش", command=lambda: run_context_action(menu, "<<Cut>>"))
    menu.add_command(label="کپی", command=lambda: run_context_action(menu, "<<Copy>>"))
    menu.add_command(label="چسباندن", command=lambda: run_context_action(menu, "<<Paste>>"))
    menu.add_command(label="انتخاب همه", command=lambda: run_context_action(menu, "<<SelectAll>>"))

    window.bind_all("<Button-3>", lambda event: show_edit_context_menu(event, menu), add="+")
    window.bind_all("<Control-Insert>", lambda event: event.widget.event_generate("<<Copy>>") if widget_supports_clipboard(event.widget) else None, add="+")
    window.bind_all("<Shift-Insert>", lambda event: event.widget.event_generate("<<Paste>>") if widget_supports_clipboard(event.widget) else None, add="+")
    window.bind_all("<Shift-Delete>", lambda event: event.widget.event_generate("<<Cut>>") if widget_supports_clipboard(event.widget) else None, add="+")

    def handle_physical_shortcut(event):
        if not widget_supports_clipboard(event.widget):
            return
        ctrl_pressed = bool(event.state & 0x4)
        if not ctrl_pressed:
            return
        keycode_map = {
            65: "<<SelectAll>>",
            67: "<<Copy>>",
            86: "<<Paste>>",
            88: "<<Cut>>",
            89: "<<Redo>>",
            90: "<<Undo>>",
        }
        action = keycode_map.get(event.keycode)
        if not action:
            return
        event.widget.event_generate(action)
        return "break"

    window.bind_all("<Control-KeyPress>", handle_physical_shortcut, add="+")


def bind_primary_action(window, callback: Callable[[], None]):
    def handle_return(event):
        widget_class = event.widget.winfo_class()
        if widget_class == "Text":
            return
        if widget_class in {"TCombobox", "Combobox"}:
            return
        callback()
        return "break"

    window.bind("<Return>", handle_return, add="+")
    window.bind("<KP_Enter>", handle_return, add="+")


def parse_jalali_datetime(value: str) -> datetime | None:
    if not value:
        return None
    formats = ["%Y/%m/%d %H:%M", "%Y-%m-%d %H:%M", "%Y/%m/%d %H:%M:%S", "%Y-%m-%d %H:%M:%S"]
    for fmt in formats:
        try:
            jd = jdatetime.datetime.strptime(value, fmt)
            return jd.togregorian()
        except ValueError:
            continue
    return None


def normalize_jalali_date(value: str) -> str:
    if not value:
        return ""
    for fmt in ["%Y/%m/%d", "%Y-%m-%d"]:
        try:
            return jdatetime.datetime.strptime(value, fmt).strftime("%Y/%m/%d")
        except ValueError:
            continue
    return value


def format_price_input(raw: str) -> str:
    digits = "".join(ch for ch in str(raw) if ch.isdigit())
    if not digits:
        return ""
    return f"{int(digits):,}"


def status_display(status: str) -> str:
    emoji = STATUS_META.get(status, {}).get("emoji", "▪")
    return f"{emoji}{emoji}{emoji}{emoji}{emoji} {status} {emoji}{emoji}{emoji}{emoji}{emoji}"


def category_display(title: str) -> str:
    cleaned = (title or "").strip()
    if not cleaned:
        return ""
    return f"  {cleaned}  "


def soften_hex_color(value: str, factor: float = 0.72) -> str:
    raw = (value or "").strip().lstrip("#")
    if len(raw) != 6:
        return "#eef2f3"
    try:
        red = int(raw[0:2], 16)
        green = int(raw[2:4], 16)
        blue = int(raw[4:6], 16)
    except ValueError:
        return "#eef2f3"
    red = int(red + (255 - red) * factor)
    green = int(green + (255 - green) * factor)
    blue = int(blue + (255 - blue) * factor)
    return f"#{red:02x}{green:02x}{blue:02x}"


@dataclass
class ProfileContext:
    profile_id: str
    title: str
    root: Path


class CsvRepository:
    def __init__(self, base_dir: Path):
        self.base_dir = base_dir
        self.base_dir.mkdir(parents=True, exist_ok=True)

    def csv_path(self, filename: str) -> Path:
        return self.base_dir / filename

    def read_rows(self, filename: str, fields: list[str]) -> list[dict]:
        path = self.csv_path(filename)
        if not path.exists():
            self.write_rows(filename, fields, [])
            return []
        with path.open("r", newline="", encoding="utf-8-sig") as handle:
            reader = csv.DictReader(handle)
            rows = []
            for row in reader:
                rows.append({field: repair_text(row.get(field, "")) for field in fields})
            return rows

    def write_rows(self, filename: str, fields: list[str], rows: list[dict]) -> None:
        path = self.csv_path(filename)
        with path.open("w", newline="", encoding="utf-8-sig") as handle:
            writer = csv.DictWriter(handle, fieldnames=fields)
            writer.writeheader()
            for row in rows:
                writer.writerow({field: repair_text(row.get(field, "")) for field in fields})


class ProfileStore:
    def __init__(self):
        self.windows_user = safe_slug(getpass.getuser())
        self.user_root = DATA_ROOT / "windows_users" / self.windows_user
        self.repo = CsvRepository(self.user_root)
        self.profiles = self.repo.read_rows("profiles.csv", PROFILE_FIELDS)

    def list_profiles(self) -> list[dict]:
        self.profiles = self.repo.read_rows("profiles.csv", PROFILE_FIELDS)
        return self.profiles

    def create_profile(self, title: str, username: str, password: str) -> ProfileContext:
        profile_id = uuid.uuid4().hex
        profile = {"id": profile_id, "title": title.strip(), "username": username.strip(), "password": password.strip(), "created_at": now_text()}
        self.profiles.append(profile)
        self.repo.write_rows("profiles.csv", PROFILE_FIELDS, self.profiles)
        profile_root = self.user_root / f"{safe_slug(title)}_{profile_id[:6]}"
        profile_root.mkdir(parents=True, exist_ok=True)
        data_repo = CsvRepository(profile_root)
        data_repo.write_rows(
            "categories.csv",
            CATEGORY_FIELDS,
            [{"id": uuid.uuid4().hex, "title": item, "color": DEFAULT_CATEGORY_COLORS[index % len(DEFAULT_CATEGORY_COLORS)]} for index, item in enumerate(DEFAULT_CATEGORIES)],
        )
        data_repo.write_rows("pipelines.csv", PIPELINE_FIELDS, [{"id": uuid.uuid4().hex, "title": item} for item in DEFAULT_PIPELINES])
        data_repo.write_rows("deal_types.csv", DEAL_TYPE_FIELDS, [{"id": uuid.uuid4().hex, "title": item} for item in DEFAULT_DEAL_TYPES])
        data_repo.write_rows("products.csv", PRODUCT_FIELDS, [])
        data_repo.write_rows("customers.csv", CUSTOMER_FIELDS, [])
        data_repo.write_rows("deals.csv", DEAL_FIELDS, [])
        return ProfileContext(profile_id=profile_id, title=title.strip(), root=profile_root)

    def resolve_profile(self, profile_id: str) -> ProfileContext | None:
        for profile in self.list_profiles():
            if profile["id"] == profile_id:
                matches = list(self.user_root.glob(f"*_{profile_id[:6]}"))
                if matches:
                    return ProfileContext(profile_id=profile_id, title=profile["title"], root=matches[0])
        return None

    def delete_profile(self, profile_id: str) -> None:
        self.profiles = [profile for profile in self.list_profiles() if profile["id"] != profile_id]
        self.repo.write_rows("profiles.csv", PROFILE_FIELDS, self.profiles)
        for path in self.user_root.glob(f"*_{profile_id[:6]}"):
            if path.exists():
                shutil.rmtree(path, ignore_errors=True)


class CrmStore:
    def __init__(self, profile: ProfileContext):
        self.profile = profile
        self.repo = CsvRepository(profile.root)
        self.ensure_defaults()

    def ensure_defaults(self) -> None:
        if not self.repo.csv_path("categories.csv").exists():
            self.repo.write_rows(
                "categories.csv",
                CATEGORY_FIELDS,
                [{"id": uuid.uuid4().hex, "title": item, "color": DEFAULT_CATEGORY_COLORS[index % len(DEFAULT_CATEGORY_COLORS)]} for index, item in enumerate(DEFAULT_CATEGORIES)],
            )
        if not self.repo.csv_path("pipelines.csv").exists():
            self.repo.write_rows("pipelines.csv", PIPELINE_FIELDS, [{"id": uuid.uuid4().hex, "title": item} for item in DEFAULT_PIPELINES])
        if not self.repo.csv_path("deal_types.csv").exists():
            self.repo.write_rows("deal_types.csv", DEAL_TYPE_FIELDS, [{"id": uuid.uuid4().hex, "title": item} for item in DEFAULT_DEAL_TYPES])
        if not self.repo.csv_path("products.csv").exists():
            self.repo.write_rows("products.csv", PRODUCT_FIELDS, [])
        if not self.repo.csv_path("customers.csv").exists():
            self.repo.write_rows("customers.csv", CUSTOMER_FIELDS, [])
        if not self.repo.csv_path("deals.csv").exists():
            self.repo.write_rows("deals.csv", DEAL_FIELDS, [])

    def customers(self) -> list[dict]:
        return self.repo.read_rows("customers.csv", CUSTOMER_FIELDS)

    def save_customers(self, rows: list[dict]) -> None:
        self.repo.write_rows("customers.csv", CUSTOMER_FIELDS, rows)

    def deals(self) -> list[dict]:
        return self.repo.read_rows("deals.csv", DEAL_FIELDS)

    def save_deals(self, rows: list[dict]) -> None:
        self.repo.write_rows("deals.csv", DEAL_FIELDS, rows)

    def categories(self) -> list[dict]:
        rows = self.repo.read_rows("categories.csv", CATEGORY_FIELDS)
        if not rows:
            return [{"id": uuid.uuid4().hex, "title": item, "color": DEFAULT_CATEGORY_COLORS[index % len(DEFAULT_CATEGORY_COLORS)]} for index, item in enumerate(DEFAULT_CATEGORIES)]
        for index, row in enumerate(rows):
            if not row.get("color"):
                row["color"] = DEFAULT_CATEGORY_COLORS[index % len(DEFAULT_CATEGORY_COLORS)]
        return rows

    def save_categories(self, rows: list[dict]) -> None:
        self.repo.write_rows("categories.csv", CATEGORY_FIELDS, rows)

    def pipelines(self) -> list[dict]:
        rows = self.repo.read_rows("pipelines.csv", PIPELINE_FIELDS)
        return rows or [{"id": uuid.uuid4().hex, "title": item} for item in DEFAULT_PIPELINES]

    def save_pipelines(self, rows: list[dict]) -> None:
        self.repo.write_rows("pipelines.csv", PIPELINE_FIELDS, rows)

    def deal_types(self) -> list[dict]:
        rows = self.repo.read_rows("deal_types.csv", DEAL_TYPE_FIELDS)
        return rows or [{"id": uuid.uuid4().hex, "title": item} for item in DEFAULT_DEAL_TYPES]

    def save_deal_types(self, rows: list[dict]) -> None:
        self.repo.write_rows("deal_types.csv", DEAL_TYPE_FIELDS, rows)

    def products(self) -> list[dict]:
        return self.repo.read_rows("products.csv", PRODUCT_FIELDS)

    def save_products(self, rows: list[dict]) -> None:
        self.repo.write_rows("products.csv", PRODUCT_FIELDS, rows)


class JalaliDatePickerDialog(tb.Toplevel):
    def __init__(self, master, initial_value: str = "", reminder_counter: Callable[[int, int], dict[int, int]] | None = None):
        super().__init__(master)
        self.result = None
        self.reminder_counter = reminder_counter
        self.title("انتخاب تاریخ شمسی")
        self.geometry("460x760")
        self.resizable(False, False)
        self.transient(master)
        self.grab_set()
        bind_edit_shortcuts(self)

        initial = parse_jalali_datetime(f"{normalize_jalali_date(initial_value)} 00:00") if initial_value else None
        if initial:
            jalali = jdatetime.datetime.fromgregorian(datetime=initial)
        else:
            jalali = jdatetime.datetime.now()

        wrapper = tb.Frame(self, padding=16)
        wrapper.pack(fill=BOTH, expand=True)

        selectors = tb.Frame(wrapper)
        selectors.pack(fill=X, pady=(0, 14))
        self.year_var = tk.IntVar(value=jalali.year)
        self.month_var = tk.IntVar(value=jalali.month)
        self.day_var = tk.IntVar(value=jalali.day)

        year_col = tb.Frame(selectors)
        year_col.pack(side=RIGHT, fill=X, expand=True, padx=4)
        rtl_label(year_col, "سال").pack(fill=X, pady=(0, 4))
        self.year_combo = ttk.Combobox(year_col, state="readonly", justify="center", values=list(range(jalali.year - 5, jalali.year + 6)), textvariable=self.year_var)
        self.year_combo.pack(fill=X)
        self.year_combo.bind("<<ComboboxSelected>>", lambda _event: self._normalize_day())

        month_col = tb.Frame(selectors)
        month_col.pack(side=RIGHT, fill=X, expand=True, padx=4)
        rtl_label(month_col, "ماه").pack(fill=X, pady=(0, 4))
        self.month_combo = ttk.Combobox(month_col, state="readonly", justify="center", values=list(range(1, 13)), textvariable=self.month_var)
        self.month_combo.pack(fill=X)
        self.month_combo.bind("<<ComboboxSelected>>", lambda _event: self._normalize_day())

        self.month_label = rtl_label(wrapper, "", font=("Tahoma", 11, "bold"))
        self.month_label.pack(fill=X, pady=(8, 10))
        self.month_summary_label = rtl_label(wrapper, "", bootstyle="info")
        self.month_summary_label.pack(fill=X, pady=(0, 6))
        self.preview_label = rtl_label(wrapper, "", bootstyle="secondary")
        self.preview_label.pack(fill=X)
        self.days_frame = tb.Frame(wrapper)
        self.days_frame.pack(fill=BOTH, expand=True, pady=(10, 0))

        actions = tb.Frame(wrapper)
        actions.pack(fill=X, pady=(18, 0))
        tb.Button(actions, text="تایید", bootstyle="success", command=self.save).pack(side=RIGHT, padx=4)
        tb.Button(actions, text="انصراف", bootstyle="secondary", command=self.destroy).pack(side=RIGHT, padx=4)

        self._normalize_day()
        self.wait_window(self)

    def _normalize_day(self):
        year = int(self.year_var.get())
        month = int(self.month_var.get())
        max_day = jdatetime.j_days_in_month[month - 1]
        if month == 12 and not jdatetime.date(year, 1, 1).isleap():
            max_day = 29
        if self.day_var.get() > max_day:
            self.day_var.set(max_day)
        reminder_counts = self.reminder_counter(year, month) if self.reminder_counter else {}
        self.month_label.configure(text=f"ماه انتخابی: {month} - {PERSIAN_MONTHS[month - 1]}")
        self.month_summary_label.configure(text=f"جمع یادآورهای این ماه: {sum(reminder_counts.values())}")
        self.preview_label.configure(text=f"{year:04d}/{month:02d}/{int(self.day_var.get()):02d}")
        for item in self.days_frame.winfo_children():
            item.destroy()
        weekday_headers = ["ش", "ی", "د", "س", "چ", "پ", "ج"]
        for column, label in enumerate(weekday_headers):
            header = rtl_label(self.days_frame, label, font=("Tahoma", 10, "bold"))
            header.grid(row=0, column=column, sticky="nsew", padx=3, pady=(0, 6))
            self.days_frame.columnconfigure(column, weight=1)
        for day in range(1, max_day + 1):
            row = ((day - 1) // 7) + 1
            column = (day - 1) % 7
            cell = tb.Frame(self.days_frame, padding=(4, 6), bootstyle="light")
            cell.grid(row=row, column=column, sticky="nsew", padx=3, pady=3)
            cell.configure(height=78, width=58)
            cell.grid_propagate(False)
            style = "success" if day == self.day_var.get() else "light-outline"
            day_button = tk.Button(
                cell,
                text=str(day),
                command=lambda d=day: self.select_day(d),
                relief="flat",
                bd=0,
                cursor="hand2",
                font=("Tahoma", 10, "bold"),
                fg="#111111",
                activeforeground="#111111",
            )
            if day == self.day_var.get():
                day_button.configure(bg="#20c997", activebackground="#18b18b")
            else:
                day_button.configure(bg="#000000", activebackground="#232527", highlightbackground="#242424")
            day_button.pack(fill=X)
            count = reminder_counts.get(day, 0)
            rtl_label(cell, f"{count} یادآور" if count else "", font=("Tahoma", 8), bootstyle="secondary").pack(fill=X, pady=(6, 0))

    def save(self):
        self.result = f"{int(self.year_var.get()):04d}/{int(self.month_var.get()):02d}/{int(self.day_var.get()):02d}"
        self.destroy()

    def select_day(self, day: int):
        self.day_var.set(day)
        self.result = f"{int(self.year_var.get()):04d}/{int(self.month_var.get()):02d}/{int(self.day_var.get()):02d}"
        self.destroy()


class JalaliDateField(tb.Frame):
    def __init__(
        self,
        master,
        initial_value: str = "",
        allow_empty: bool = False,
        reminder_counter: Callable[[int, int], dict[int, int]] | None = None,
        on_change: Callable[[], None] | None = None,
    ):
        super().__init__(master)
        self.allow_empty = allow_empty
        self.reminder_counter = reminder_counter
        self.on_change = on_change
        self._change_after_id = None
        initial = normalize_jalali_date(initial_value)
        if not initial and not allow_empty:
            initial = jdatetime.date.today().strftime("%Y/%m/%d")
        self.value_var = tk.StringVar(value=initial)
        self.value_var.trace_add("write", self._on_value_changed)
        self.entry = tb.Entry(self, textvariable=self.value_var, justify="center")
        self.entry.pack(side=RIGHT, fill=X, expand=True)
        if allow_empty:
            tb.Button(self, text="پاک", bootstyle="secondary-outline", command=lambda: self.value_var.set(""), width=6).pack(side=RIGHT, padx=(0, 6))
        tb.Button(self, text="تقویم", bootstyle="info-outline", command=self.open_picker, width=9).pack(side=RIGHT, padx=(0, 6))

    def _on_value_changed(self, *_args):
        if not self.on_change:
            return
        if self._change_after_id:
            self.after_cancel(self._change_after_id)
        self._change_after_id = self.after(10, self._emit_change)

    def _emit_change(self):
        self._change_after_id = None
        if self.on_change:
            self.on_change()

    def open_picker(self):
        dialog = JalaliDatePickerDialog(self, self.value_var.get(), reminder_counter=self.reminder_counter)
        if dialog.result:
            self.value_var.set(dialog.result)

    def get(self) -> str:
        return self.value_var.get().strip()

    def set(self, value: str):
        self.value_var.set(normalize_jalali_date(value))


class SimpleLoginDialog(tb.Toplevel):
    def __init__(self, master, title: str, default_username: str = ""):
        super().__init__(master)
        self.result = None
        self.title(title)
        self.geometry("340x220")
        self.transient(master)
        self.grab_set()
        bind_edit_shortcuts(self)
        wrapper = tb.Frame(self, padding=16)
        wrapper.pack(fill=BOTH, expand=True)
        rtl_label(wrapper, "یوزرنیم").pack(fill=X, pady=(0, 4))
        self.username = tb.Entry(wrapper, justify="right")
        self.username.pack(fill=X, pady=(0, 10))
        self.username.insert(0, default_username)
        rtl_label(wrapper, "پسورد").pack(fill=X, pady=(0, 4))
        self.password = tb.Entry(wrapper, justify="right", show="*")
        self.password.pack(fill=X, pady=(0, 12))
        actions = tb.Frame(wrapper)
        actions.pack(fill=X)
        tb.Button(actions, text="ورود", bootstyle="success", command=self.submit).pack(side=RIGHT, padx=4)
        tb.Button(actions, text="انصراف", bootstyle="secondary", command=self.destroy).pack(side=RIGHT, padx=4)
        bind_primary_action(self, self.submit)
        self.wait_window(self)

    def submit(self):
        self.result = {"username": self.username.get().strip(), "password": self.password.get().strip()}
        self.destroy()


class SearchableCombobox(tb.Entry):
    def __init__(self, master, values: list[str], **kwargs):
        super().__init__(master, **kwargs)
        self.all_values = list(values)
        self.filtered_values = list(values)
        self.popup: tk.Toplevel | None = None
        self.listbox: tk.Listbox | None = None
        self.popup_visible = False
        self.bind("<KeyRelease>", self._filter_values)
        self.bind("<FocusIn>", lambda _event: self._filter_values())
        self.bind("<FocusOut>", lambda _event: self.after(150, self.hide_popup))
        self.bind("<Down>", self.focus_listbox)

    def set_values(self, values: list[str]):
        self.all_values = list(values)
        self.filtered_values = list(values)
        self._fill_listbox(self.filtered_values)

    def set(self, value: str):
        self.delete(0, END)
        self.insert(0, value)
        self.hide_popup(force=True)

    def ensure_popup(self):
        if self.popup and self.popup.winfo_exists():
            return
        self.popup = tk.Toplevel(self)
        self.popup.withdraw()
        self.popup.overrideredirect(True)
        self.popup.attributes("-topmost", True)
        wrapper = tb.Frame(self.popup, bootstyle="light")
        wrapper.pack(fill=BOTH, expand=True)
        scrollbar = tb.Scrollbar(wrapper, orient="vertical")
        scrollbar.pack(side=LEFT, fill=Y)
        self.listbox = tk.Listbox(wrapper, height=6, font=("Tahoma", 10), activestyle="none", justify="right", yscrollcommand=scrollbar.set)
        self.listbox.pack(side=RIGHT, fill=BOTH, expand=True)
        scrollbar.configure(command=self.listbox.yview)
        self.listbox.bind("<ButtonRelease-1>", self.select_from_list)
        self.listbox.bind("<Double-Button-1>", self.select_from_list)
        self.listbox.bind("<Return>", self.select_from_list)
        self.listbox.bind("<FocusOut>", lambda _event: self.after(150, self.hide_popup))

    def _fill_listbox(self, values: list[str]):
        self.ensure_popup()
        self.listbox.delete(0, END)
        for item in values:
            self.listbox.insert(END, item)

    def _filter_values(self, _event=None):
        query = normalize_keyboard_text(self.get().strip())
        if not query:
            self.filtered_values = list(self.all_values)
        else:
            self.filtered_values = [item for item in self.all_values if query in normalize_keyboard_text(item)]
        self._fill_listbox(self.filtered_values)
        if self.filtered_values and query:
            self.show_popup()
        elif not query:
            self.hide_popup(force=True)
        else:
            self.hide_popup(force=True)

    def show_popup(self):
        self.ensure_popup()
        x = self.winfo_rootx()
        y = self.winfo_rooty() + self.winfo_height() + 2
        width = self.winfo_width()
        self.popup.geometry(f"{width}x180+{x}+{y}")
        self.popup.deiconify()
        self.popup_visible = True

    def hide_popup(self, force: bool = False):
        if not force:
            focus_widget = self.focus_get()
            if focus_widget in {self, self.listbox}:
                return
        if self.popup and self.popup.winfo_exists():
            self.popup.withdraw()
        self.popup_visible = False

    def focus_listbox(self, _event=None):
        if not self.filtered_values:
            return "break"
        self.show_popup()
        self.listbox.focus_set()
        self.listbox.selection_clear(0, END)
        self.listbox.selection_set(0)
        self.listbox.activate(0)
        return "break"

    def select_from_list(self, _event=None):
        if not self.listbox:
            return "break"
        selection = self.listbox.curselection()
        if not selection:
            return "break"
        value = self.listbox.get(selection[0])
        self.set(value)
        self.focus_set()
        self.icursor(END)
        return "break"


class ProfileDialog(tb.Toplevel):
    def __init__(self, master, store: ProfileStore):
        super().__init__(master)
        self.store = store
        self.result: ProfileContext | None = None
        self.title("انتخاب پروفایل")
        self.geometry("820x650")
        self.minsize(720, 520)
        self.resizable(True, True)
        self.grab_set()
        self.lift()
        self.focus_force()
        bind_edit_shortcuts(self)

        wrapper = tb.Frame(self, padding=18)
        wrapper.pack(fill=BOTH, expand=True)
        rtl_label(wrapper, f"کاربر ویندوز: {getpass.getuser()}", font=("Tahoma", 11, "bold")).pack(fill=X)
        rtl_label(
            wrapper,
            "برای شروع یک پروفایل اختصاصی بسازید یا یکی از پروفایل‌های قبلی را باز کنید.",
            bootstyle="secondary",
        ).pack(fill=X, pady=(8, 16))

        self.tree = ttk.Treeview(wrapper, columns=("title", "created"), show="headings", height=8)
        self.tree.heading("title", text="نام پروفایل")
        self.tree.heading("created", text="تاریخ ایجاد")
        self.tree.column("title", width=250, anchor="e")
        self.tree.column("created", width=180, anchor="center")
        self.tree.pack(fill=BOTH, expand=True)
        self.tree.bind("<Double-1>", lambda _event: self.open_selected())

        form = tb.LabelFrame(wrapper, text="ساخت پروفایل جدید")
        form.pack(fill=X, pady=14)
        form_body = tb.Frame(form, padding=12)
        form_body.pack(fill=BOTH, expand=True)
        rtl_label(form_body, "نام پروفایل").pack(fill=X, pady=(0, 4))
        self.profile_name = tb.Entry(form_body, justify="right")
        self.profile_name.pack(fill=X, pady=(0, 8))
        rtl_label(form_body, "یوزرنیم").pack(fill=X, pady=(0, 4))
        self.profile_username = tb.Entry(form_body, justify="right")
        self.profile_username.pack(fill=X, pady=(0, 8))
        rtl_label(form_body, "پسورد").pack(fill=X, pady=(0, 4))
        self.profile_password = tb.Entry(form_body, justify="right", show="*")
        self.profile_password.pack(fill=X)

        actions = tb.Frame(wrapper)
        actions.pack(fill=X)
        tb.Button(actions, text="باز کردن پروفایل", bootstyle="success", command=self.open_selected).pack(side=RIGHT, padx=4)
        tb.Button(actions, text="ایجاد پروفایل", bootstyle="primary", command=self.create_profile).pack(side=RIGHT, padx=4)
        tb.Button(actions, text="خروج", bootstyle="secondary", command=self.cancel).pack(side=LEFT, padx=4)

        tb.Button(actions, text="حذف پروفایل", bootstyle="danger-outline", command=self.delete_selected).pack(side=RIGHT, padx=4)
        bind_primary_action(self, self.handle_primary_enter)
        self.refresh_profiles()
        self.wait_window(self)

    def handle_primary_enter(self):
        focused = self.focus_get()
        if focused == self.tree or self.tree.selection():
            self.open_selected()
            return
        self.create_profile()

    def refresh_profiles(self):
        for item in self.tree.get_children():
            self.tree.delete(item)
        for profile in self.store.list_profiles():
            self.tree.insert("", END, iid=profile["id"], values=(profile["title"], profile["created_at"]))

    def open_selected(self):
        selected = self.tree.selection()
        if not selected:
            Messagebox.show_warning("یک پروفایل را انتخاب کنید.", "انتخاب پروفایل", parent=self)
            return
        profile = next((item for item in self.store.list_profiles() if item["id"] == selected[0]), None)
        if not profile:
            return
        username = Messagebox.okcancel("برای ورود، یوزرنیم و پسورد را در پنجره بعدی وارد کنید.", "ورود", parent=self)
        if username != "OK":
            return
        auth = SimpleLoginDialog(self, "ورود به پروفایل", default_username=profile.get("username", "admin") or "admin")
        if not auth.result:
            return
        expected_user = profile.get("username", "") or "admin"
        expected_pass = profile.get("password", "") or "admin"
        if auth.result["username"] != expected_user or auth.result["password"] != expected_pass:
            Messagebox.show_error("نام کاربری یا رمز عبور اشتباه است.", "ورود", parent=self)
            return
        self.result = self.store.resolve_profile(selected[0])
        self.destroy()

    def create_profile(self):
        title = self.profile_name.get().strip()
        username = self.profile_username.get().strip()
        password = self.profile_password.get().strip()
        if not title or not username or not password:
            Messagebox.show_error("نام پروفایل، یوزرنیم و پسورد الزامی است.", "خطا", parent=self)
            return
        self.result = self.store.create_profile(title, username, password)
        self.destroy()

    def delete_selected(self):
        selected = self.tree.selection()
        if not selected:
            Messagebox.show_warning("یک پروفایل را انتخاب کنید.", "حذف پروفایل", parent=self)
            return
        profile = next((item for item in self.store.list_profiles() if item["id"] == selected[0]), None)
        if not profile:
            return
        auth = SimpleLoginDialog(self, "تایید حذف پروفایل", default_username=profile.get("username", "admin") or "admin")
        if not auth.result:
            return
        expected_user = profile.get("username", "") or "admin"
        expected_pass = profile.get("password", "") or "admin"
        if auth.result["username"] != expected_user or auth.result["password"] != expected_pass:
            Messagebox.show_error("نام کاربری یا رمز عبور اشتباه است.", "حذف پروفایل", parent=self)
            return
        confirm = Messagebox.okcancel(f"پروفایل «{profile['title']}» حذف شود؟", "تایید", parent=self)
        if confirm != "OK":
            return
        self.store.delete_profile(profile["id"])
        self.refresh_profiles()

    def cancel(self):
        self.result = None
        self.destroy()


class CustomerDialog(tb.Toplevel):
    def __init__(self, master, customer: dict | None = None, deals: list[dict] | None = None, open_deal_callback: Callable[[str], None] | None = None, existing_mobiles: list[str] | None = None):
        super().__init__(master)
        self.result = None
        self.open_deal_callback = open_deal_callback
        self.existing_mobiles = set(existing_mobiles or [])
        self.original_mobile = customer.get("mobile", "") if customer else ""
        self.title("ثبت مشتری" if not customer else "ویرایش مشتری")
        self.geometry("874x620" if customer else "706x360")
        self.resizable(False, False)
        self.transient(master)
        self.grab_set()

        wrapper = tb.Frame(self, padding=16)
        wrapper.pack(fill=BOTH, expand=True)
        self.entries: dict[str, object] = {}
        for key, title in [("name", "نام"), ("mobile", "شماره موبایل")]:
            rtl_label(wrapper, title).pack(fill=X, pady=(0, 4))
            entry = tb.Entry(wrapper, justify="right")
            entry.pack(fill=X, pady=(0, 10))
            self.entries[key] = entry
            if key == "mobile":
                self.mobile_error = rtl_label(wrapper, "", bootstyle="danger")
                self.mobile_error.pack(fill=X, pady=(0, 8))

        rtl_label(wrapper, "سطح درآمد").pack(fill=X, pady=(0, 4))
        income_combo = ttk.Combobox(wrapper, values=INCOME_LEVELS, justify="right", state="readonly")
        income_combo.pack(fill=X, pady=(0, 10))
        income_combo.set(INCOME_LEVELS[1])
        self.entries["income_level"] = income_combo

        rtl_label(wrapper, "تاریخ تولد").pack(fill=X, pady=(0, 4))
        birthdate = JalaliDateField(wrapper)
        birthdate.pack(fill=X, pady=(0, 12))
        self.entries["birthdate"] = birthdate

        actions = tb.Frame(wrapper)
        actions.pack(fill=X, pady=(12, 0))
        tb.Button(actions, text="ذخیره", bootstyle="success", command=self.save).pack(side=RIGHT, padx=4)
        tb.Button(actions, text="انصراف", bootstyle="secondary", command=self.destroy).pack(side=RIGHT, padx=4)
        bind_primary_action(self, self.save)

        if customer:
            deals_box = tb.LabelFrame(wrapper, text="معاملات این مشتری")
            deals_box.pack(fill=BOTH, expand=True, pady=(14, 0))
            deals_body = tb.Frame(deals_box, padding=10)
            deals_body.pack(fill=BOTH, expand=True)
            self.deals_tree = ttk.Treeview(deals_body, columns=("title", "status", "price", "reminder"), show="headings", height=8)
            for key, title, width in [("title", "عنوان", 180), ("status", "وضعیت", 110), ("price", "قیمت", 100), ("reminder", "یادآور", 120)]:
                self.deals_tree.heading(key, text=title)
                self.deals_tree.column(key, width=width, anchor="e")
            self.deals_tree.pack(fill=BOTH, expand=True)
            self.deals_tree.bind("<Double-1>", self.open_deal_from_tree)
            for deal in deals or []:
                status_label = status_display(deal["status"])
                self.deals_tree.insert("", END, iid=deal["id"], values=(deal["title"], status_label, deal["sale_price"], deal["reminder_at"]))

        if customer:
            self.entries["name"].insert(0, customer.get("name", ""))
            self.entries["mobile"].insert(0, customer.get("mobile", ""))
            self.entries["income_level"].set(customer.get("income_level", INCOME_LEVELS[1]))
            birthdate.set(customer.get("birthdate", ""))

        self.wait_window(self)

    def save(self):
        name = self.entries["name"].get().strip()
        mobile = self.entries["mobile"].get().strip()
        self.mobile_error.configure(text="")
        if not name or not mobile:
            Messagebox.show_error("نام و شماره موبایل الزامی است.", "خطا", parent=self)
            return
        if mobile != self.original_mobile and mobile in self.existing_mobiles:
            self.mobile_error.configure(text="این شماره تکراری است.")
            return
        self.result = {
            "name": name,
            "mobile": mobile,
            "income_level": self.entries["income_level"].get().strip(),
            "birthdate": self.entries["birthdate"].get().strip(),
            "birthday_notified_for": "",
        }
        self.destroy()

    def open_deal_from_tree(self, _event=None):
        if not self.open_deal_callback or not hasattr(self, "deals_tree"):
            return
        selected = self.deals_tree.selection()
        if selected:
            self.open_deal_callback(selected[0])


class TextListManager(tb.LabelFrame):
    def __init__(self, master, title: str, initial_rows: list[dict], save_callback: Callable[[list[dict]], None]):
        super().__init__(master, text=title)
        self.rows = [dict(row) for row in initial_rows]
        self.save_callback = save_callback

        body = tb.Frame(self, padding=12)
        body.pack(fill=BOTH, expand=True)

        self.tree = ttk.Treeview(body, columns=("title",), show="headings", height=8)
        self.tree.heading("title", text=title)
        self.tree.column("title", anchor="e", width=220)
        self.tree.pack(fill=BOTH, expand=True)

        self.entry = tb.Entry(body, justify="right")
        self.entry.pack(fill=X, pady=8)

        actions = tb.Frame(body)
        actions.pack(fill=X)
        tb.Button(actions, text="افزودن", bootstyle="success-outline", command=self.add_item).pack(side=RIGHT, padx=3)
        tb.Button(actions, text="ویرایش", bootstyle="warning-outline", command=self.edit_item).pack(side=RIGHT, padx=3)
        tb.Button(actions, text="حذف", bootstyle="danger-outline", command=self.delete_item).pack(side=RIGHT, padx=3)

        self.refresh()

    def refresh(self):
        for item in self.tree.get_children():
            self.tree.delete(item)
        for row in self.rows:
            self.tree.insert("", END, iid=row["id"], values=(row["title"],))

    def add_item(self):
        title = self.entry.get().strip()
        if not title:
            return
        self.rows.append({"id": uuid.uuid4().hex, "title": title})
        self.entry.delete(0, END)
        self.refresh()
        self.save_callback(self.rows)

    def edit_item(self):
        selected = self.tree.selection()
        if not selected:
            return
        row = next((item for item in self.rows if item["id"] == selected[0]), None)
        if not row:
            return
        row["title"] = self.entry.get().strip() or row["title"]
        self.refresh()
        self.save_callback(self.rows)

    def delete_item(self):
        selected = self.tree.selection()
        if not selected:
            return
        self.rows = [item for item in self.rows if item["id"] != selected[0]]
        self.refresh()
        self.save_callback(self.rows)


class CategoryManager(tb.LabelFrame):
    def __init__(self, master, rows: list[dict], save_callback: Callable[[list[dict]], None]):
        super().__init__(master, text="دسته‌بندی‌های معامله")
        self.rows = [dict(row) for row in rows]
        self.save_callback = save_callback
        body = tb.Frame(self, padding=12)
        body.pack(fill=BOTH, expand=True)
        self.tree = ttk.Treeview(body, columns=("title", "color"), show="headings", height=8)
        self.tree.heading("title", text="عنوان")
        self.tree.heading("color", text="رنگ")
        self.tree.column("title", anchor="e", width=160)
        self.tree.column("color", anchor="center", width=90)
        self.tree.pack(fill=BOTH, expand=True)
        form = tb.Frame(body)
        form.pack(fill=X, pady=8)
        self.entry = tb.Entry(form, justify="right")
        self.entry.pack(side=RIGHT, fill=X, expand=True)
        self.color_preview = tb.Label(form, text="      ", background="#457b9d")
        self.color_preview.pack(side=RIGHT, padx=6)
        self.current_color = "#457b9d"
        tb.Button(form, text="انتخاب رنگ", bootstyle="info-outline", command=self.pick_color).pack(side=RIGHT, padx=4)
        actions = tb.Frame(body)
        actions.pack(fill=X)
        tb.Button(actions, text="افزودن", bootstyle="success-outline", command=self.add_item).pack(side=RIGHT, padx=3)
        tb.Button(actions, text="ویرایش", bootstyle="warning-outline", command=self.edit_item).pack(side=RIGHT, padx=3)
        tb.Button(actions, text="حذف", bootstyle="danger-outline", command=self.delete_item).pack(side=RIGHT, padx=3)
        self.tree.bind("<<TreeviewSelect>>", self.load_selected)
        self.refresh()

    def pick_color(self):
        chosen = colorchooser.askcolor(color=self.current_color, parent=self)
        if chosen and chosen[1]:
            self.current_color = chosen[1]
            self.color_preview.configure(background=self.current_color)

    def load_selected(self, _event=None):
        selected = self.tree.selection()
        if not selected:
            return
        row = next((item for item in self.rows if item["id"] == selected[0]), None)
        if row:
            self.entry.delete(0, END)
            self.entry.insert(0, row["title"])
            self.current_color = row.get("color") or "#457b9d"
            self.color_preview.configure(background=self.current_color)

    def refresh(self):
        for item in self.tree.get_children():
            self.tree.delete(item)
        for row in self.rows:
            self.tree.insert("", END, iid=row["id"], values=(row["title"], row.get("color", "")))

    def add_item(self):
        title = self.entry.get().strip()
        if not title:
            return
        self.rows.append({"id": uuid.uuid4().hex, "title": title, "color": self.current_color})
        self.entry.delete(0, END)
        self.refresh()
        self.save_callback(self.rows)

    def edit_item(self):
        selected = self.tree.selection()
        if not selected:
            return
        row = next((item for item in self.rows if item["id"] == selected[0]), None)
        if not row:
            return
        row["title"] = self.entry.get().strip() or row["title"]
        row["color"] = self.current_color
        self.refresh()
        self.save_callback(self.rows)

    def delete_item(self):
        selected = self.tree.selection()
        if not selected:
            return
        self.rows = [item for item in self.rows if item["id"] != selected[0]]
        self.refresh()
        self.save_callback(self.rows)


class ProductManager(tb.LabelFrame):
    def __init__(self, master, deal_types: list[dict], products: list[dict], save_callback: Callable[[list[dict]], None]):
        super().__init__(master, text="محصولات نوع معامله")
        self.deal_types = [dict(row) for row in deal_types]
        self.rows = [dict(row) for row in products]
        self.save_callback = save_callback
        body = tb.Frame(self, padding=12)
        body.pack(fill=BOTH, expand=True)
        self.tree = ttk.Treeview(body, columns=("type", "title"), show="headings", height=8)
        self.tree.heading("type", text="نوع معامله")
        self.tree.heading("title", text="محصول")
        self.tree.column("type", anchor="e", width=130)
        self.tree.column("title", anchor="e", width=160)
        self.tree.pack(fill=BOTH, expand=True)
        rtl_label(body, "نوع معامله").pack(fill=X, pady=(8, 4))
        self.type_combo = ttk.Combobox(body, state="readonly", justify="right", values=[row["title"] for row in self.deal_types])
        self.type_combo.pack(fill=X)
        if self.deal_types:
            self.type_combo.set(self.deal_types[0]["title"])
        rtl_label(body, "نام محصول").pack(fill=X, pady=(8, 4))
        self.entry = tb.Entry(body, justify="right")
        self.entry.pack(fill=X)
        actions = tb.Frame(body)
        actions.pack(fill=X, pady=8)
        tb.Button(actions, text="افزودن", bootstyle="success-outline", command=self.add_item).pack(side=RIGHT, padx=3)
        tb.Button(actions, text="ویرایش", bootstyle="warning-outline", command=self.edit_item).pack(side=RIGHT, padx=3)
        tb.Button(actions, text="حذف", bootstyle="danger-outline", command=self.delete_item).pack(side=RIGHT, padx=3)
        self.tree.bind("<<TreeviewSelect>>", self.load_selected)
        self.refresh()

    def set_deal_types(self, deal_types: list[dict]):
        self.deal_types = [dict(row) for row in deal_types]
        values = [row["title"] for row in self.deal_types]
        self.type_combo.configure(values=values)
        if values and self.type_combo.get() not in values:
            self.type_combo.set(values[0])

    def refresh(self):
        for item in self.tree.get_children():
            self.tree.delete(item)
        for row in self.rows:
            self.tree.insert("", END, iid=row["id"], values=(row["deal_type"], row["title"]))

    def load_selected(self, _event=None):
        selected = self.tree.selection()
        if not selected:
            return
        row = next((item for item in self.rows if item["id"] == selected[0]), None)
        if row:
            self.entry.delete(0, END)
            self.entry.insert(0, row["title"])
            self.type_combo.set(row["deal_type"])

    def add_item(self):
        title = self.entry.get().strip()
        deal_type = self.type_combo.get().strip()
        if not title or not deal_type:
            return
        self.rows.append({"id": uuid.uuid4().hex, "deal_type": deal_type, "title": title})
        self.entry.delete(0, END)
        self.refresh()
        self.save_callback(self.rows)

    def edit_item(self):
        selected = self.tree.selection()
        if not selected:
            return
        row = next((item for item in self.rows if item["id"] == selected[0]), None)
        if not row:
            return
        row["title"] = self.entry.get().strip() or row["title"]
        row["deal_type"] = self.type_combo.get().strip() or row["deal_type"]
        self.refresh()
        self.save_callback(self.rows)

    def delete_item(self):
        selected = self.tree.selection()
        if not selected:
            return
        self.rows = [item for item in self.rows if item["id"] != selected[0]]
        self.refresh()
        self.save_callback(self.rows)


class NoteEditorDialog(tb.Toplevel):
    RICH_TAGS = ("bold", "italic", "underline", "heading", "link", "timestamp", "quote")

    def __init__(self, master, note: dict[str, str] | None = None):
        super().__init__(master)
        self.result = None
        self.original_note = dict(note) if note else None
        self.title("افزودن یادداشت" if note is None else "ویرایش یادداشت")
        self.geometry("760x560")
        self.minsize(680, 480)
        self.resizable(True, True)
        self.transient(master)
        self.grab_set()
        bind_edit_shortcuts(self)

        container = tb.Frame(self, padding=16)
        container.pack(fill=BOTH, expand=True)

        actions = tb.Frame(container, padding=(0, 0, 0, 10))
        actions.pack(fill=X)
        tb.Button(actions, text="ثبت", width=12, bootstyle="success", command=self.submit).pack(side=RIGHT, padx=4)
        tb.Button(actions, text="لغو", width=12, bootstyle="secondary", command=self.destroy).pack(side=RIGHT, padx=4)
        tb.Separator(container).pack(fill=X, pady=(0, 10))

        wrapper = tb.Frame(container)
        wrapper.pack(fill=BOTH, expand=True)
        rtl_label(wrapper, "متن یادداشت", font=("Tahoma", 11, "bold")).pack(fill=X, pady=(0, 8))

        toolbar_primary = tb.Frame(wrapper)
        toolbar_primary.pack(fill=X, pady=(0, 6))
        tb.Button(toolbar_primary, text="پررنگ", bootstyle="secondary-outline", command=lambda: self.toggle_tag("bold")).pack(side=RIGHT, padx=3)
        tb.Button(toolbar_primary, text="مورب", bootstyle="secondary-outline", command=lambda: self.toggle_tag("italic")).pack(side=RIGHT, padx=3)
        tb.Button(toolbar_primary, text="زیرخط", bootstyle="secondary-outline", command=lambda: self.toggle_tag("underline")).pack(side=RIGHT, padx=3)
        tb.Button(toolbar_primary, text="تیتر", bootstyle="info-outline", command=lambda: self.toggle_line_tag("heading")).pack(side=RIGHT, padx=3)
        tb.Button(toolbar_primary, text="لینک", bootstyle="info-outline", command=self.insert_link).pack(side=RIGHT, padx=3)
        tb.Button(toolbar_primary, text="زمان", bootstyle="success-outline", command=self.insert_timestamp).pack(side=RIGHT, padx=3)

        toolbar_secondary = tb.Frame(wrapper)
        toolbar_secondary.pack(fill=X, pady=(0, 8))
        tb.Button(toolbar_secondary, text="بولت", bootstyle="secondary-outline", command=lambda: self.insert_prefix("• ")).pack(side=RIGHT, padx=3)
        tb.Button(toolbar_secondary, text="چک‌لیست", bootstyle="secondary-outline", command=lambda: self.insert_prefix("☐ ")).pack(side=RIGHT, padx=3)
        tb.Button(toolbar_secondary, text="نقل‌قول", bootstyle="secondary-outline", command=self.apply_quote_block).pack(side=RIGHT, padx=3)
        tb.Button(toolbar_secondary, text="شماره", bootstyle="secondary-outline", command=lambda: self.insert_prefix("1. ")).pack(side=RIGHT, padx=3)
        tb.Button(toolbar_secondary, text="جداکننده", bootstyle="warning-outline", command=lambda: self.insert_block("\n────────────────\n")).pack(side=RIGHT, padx=3)

        editor_box = tb.Frame(wrapper, bootstyle="light", padding=10)
        editor_box.pack(fill=BOTH, expand=True)
        scroll = tb.Scrollbar(editor_box, orient="vertical")
        scroll.pack(side=LEFT, fill=Y)
        self.text = tk.Text(
            editor_box,
            wrap="word",
            undo=True,
            font=("Tahoma", 10),
            padx=12,
            pady=10,
            spacing1=3,
            spacing3=4,
            insertwidth=2,
            insertbackground="#111111",
            yscrollcommand=scroll.set,
        )
        self.text.pack(side=RIGHT, fill=BOTH, expand=True)
        scroll.configure(command=self.text.yview)
        self.configure_editor_widget(self.text)
        self.text.bind("<<Modified>>", self.on_text_modified, add="+")
        self.text.bind("<Control-b>", lambda _event: self.toggle_tag("bold") or "break", add="+")
        self.text.bind("<Control-i>", lambda _event: self.toggle_tag("italic") or "break", add="+")
        self.text.bind("<Control-u>", lambda _event: self.toggle_tag("underline") or "break", add="+")
        self.text.bind("<Control-l>", lambda _event: self.insert_link() or "break", add="+")
        self.text.bind("<KeyRelease>", self.on_editor_key_release, add="+")
        self.text.bind("<<Paste>>", self.on_editor_paste, add="+")

        bind_primary_action(self, self.submit)

        if note:
            self.load_note_content(note.get("text", ""))
        self.apply_alignment()
        self.text.focus_set()
        self.wait_window(self)

    def configure_editor_widget(self, widget: tk.Text):
        widget.tag_configure("rtl", justify="right", rmargin=14, lmargin1=14, lmargin2=14, spacing1=3, spacing3=4)
        widget.tag_configure("bold", font=("Tahoma", 10, "bold"))
        widget.tag_configure("italic", font=("Tahoma", 10, "italic"))
        widget.tag_configure("underline", underline=True)
        widget.tag_configure("heading", font=("Tahoma", 13, "bold"), spacing1=8, spacing3=6)
        widget.tag_configure("link", foreground="#0d6efd", underline=True)
        widget.tag_configure("timestamp", foreground="#198754", font=("Tahoma", 9, "bold"))
        widget.tag_configure("quote", background="#f8f9fa", foreground="#495057", lmargin1=28, lmargin2=28, rmargin=18)
        widget.tag_add("rtl", "1.0", "end")

    def on_text_modified(self, _event=None):
        try:
            if self.text.edit_modified():
                self.apply_alignment()
                self.text.edit_modified(False)
        except tk.TclError:
            pass

    def on_editor_key_release(self, event=None):
        if event and event.keysym in {"Up", "Down", "Left", "Right", "Home", "End", "Prior", "Next", "Shift_L", "Shift_R", "Control_L", "Control_R", "Alt_L", "Alt_R"}:
            return
        self.after_idle(self.normalize_editor_text)

    def on_editor_paste(self, _event=None):
        self.after_idle(self.normalize_editor_text)

    def normalize_editor_text(self):
        exported = self.export_note_payload()
        original_text = exported["text"]
        normalized_text = normalize_persian_editor_text(original_text)
        if normalized_text == original_text:
            self.apply_alignment()
            return
        insert_offset = self.index_to_offset(self.text.index("insert"))
        selected = self.selected_range()
        selection_offsets = None
        if selected:
            selection_offsets = (self.index_to_offset(selected[0]), self.index_to_offset(selected[1]))
        exported["text"] = normalized_text
        self.load_note_payload(exported)
        self.text.mark_set("insert", self.offset_to_index(min(insert_offset, len(normalized_text))))
        if selection_offsets:
            start_offset, end_offset = selection_offsets
            self.text.tag_add("sel", self.offset_to_index(min(start_offset, len(normalized_text))), self.offset_to_index(min(end_offset, len(normalized_text))))
        self.apply_alignment()

    def apply_alignment(self):
        self.text.tag_add("rtl", "1.0", "end")

    def selected_range(self) -> tuple[str, str] | None:
        try:
            start = self.text.index("sel.first")
            end = self.text.index("sel.last")
            return start, end
        except tk.TclError:
            return None

    def toggle_tag(self, tag_name: str):
        selected = self.selected_range()
        if not selected:
            return
        start, end = selected
        if tag_name in self.text.tag_names("sel.first"):
            self.text.tag_remove(tag_name, start, end)
        else:
            self.text.tag_add(tag_name, start, end)
        self.apply_alignment()

    def toggle_line_tag(self, tag_name: str):
        line_start = self.text.index("insert linestart")
        line_end = self.text.index("insert lineend")
        if tag_name in self.text.tag_names(line_start):
            self.text.tag_remove(tag_name, line_start, line_end)
        else:
            self.text.tag_add(tag_name, line_start, line_end)
        self.apply_alignment()

    def insert_prefix(self, prefix: str):
        self.text.insert("insert linestart", prefix)
        self.apply_alignment()

    def apply_quote_block(self):
        line_start = self.text.index("insert linestart")
        line_end = self.text.index("insert lineend")
        self.text.tag_add("quote", line_start, line_end)
        self.apply_alignment()

    def insert_block(self, block: str):
        self.text.insert("insert", block)
        self.apply_alignment()

    def insert_timestamp(self):
        start = self.text.index("insert")
        self.text.insert("insert", f"[{now_text()}]")
        end = self.text.index("insert")
        self.text.tag_add("timestamp", start, end)
        self.apply_alignment()

    def insert_link(self):
        selected = self.selected_range()
        url = simpledialog.askstring("لینک", "آدرس لینک را وارد کنید:", parent=self)
        if not url:
            return
        label = self.text.get(*selected) if selected else url
        visible_text = label if label != url else url.strip()
        if selected:
            self.text.delete(selected[0], selected[1])
            self.text.insert(selected[0], visible_text)
            start, end = selected[0], f"{selected[0]}+{len(visible_text)}c"
        else:
            start = self.text.index("insert")
            self.text.insert("insert", visible_text)
            end = self.text.index("insert")
        self.text.tag_add("link", start, end)
        self.apply_alignment()

    def index_to_offset(self, index: str) -> int:
        return len(self.text.get("1.0", index))

    def offset_to_index(self, offset: int) -> str:
        return f"1.0+{offset}c"

    def export_note_payload(self) -> dict:
        text = self.text.get("1.0", "end-1c")
        spans = []
        for tag_name in self.RICH_TAGS:
            ranges = self.text.tag_ranges(tag_name)
            for idx in range(0, len(ranges), 2):
                start = self.index_to_offset(str(ranges[idx]))
                end = self.index_to_offset(str(ranges[idx + 1]))
                if end > start:
                    spans.append({"tag": tag_name, "start": start, "end": end})
        return {"text": text, "spans": spans}

    def export_note_content(self) -> str:
        payload = self.export_note_payload()
        return f"{RICH_NOTE_PREFIX}{json.dumps(payload, ensure_ascii=False)}"

    def load_note_payload(self, payload: dict):
        self.text.delete("1.0", END)
        text = payload.get("text", "")
        self.text.insert("1.0", text)
        for span in payload.get("spans", []):
            tag_name = span.get("tag")
            if tag_name not in self.RICH_TAGS:
                continue
            start = self.offset_to_index(int(span.get("start", 0)))
            end = self.offset_to_index(int(span.get("end", 0)))
            self.text.tag_add(tag_name, start, end)
        self.apply_alignment()

    def load_note_content(self, value: str):
        payload = decode_rich_note(value or "")
        if payload:
            self.load_note_payload(payload)
        else:
            self.text.delete("1.0", END)
            self.text.insert("1.0", normalize_persian_editor_text(value or ""))
            self.apply_alignment()

    def submit(self):
        plain_text = self.text.get("1.0", END).strip()
        if not plain_text:
            Messagebox.show_error("متن یادداشت نمی‌تواند خالی باشد.", "یادداشت", parent=self)
            return
        serialized = self.export_note_content()
        self.result = {
            "time": (self.original_note or {}).get("time", now_text()),
            "text": serialized,
            "preview": note_preview_text(serialized, limit=120) or "-",
        }
        self.destroy()


class DealDialog(tb.Toplevel):
    def __init__(
        self,
        master,
        customers: list[dict],
        categories: list[dict],
        pipelines: list[dict],
        deal_types: list[dict],
        products: list[dict],
        existing_deals: list[dict] | None = None,
        deal: dict | None = None,
    ):
        super().__init__(master)
        self.result = None
        self.products = [dict(item) for item in products]
        self.customers = list(customers)
        self.customer_lookup = {item["name"]: item for item in customers}
        self.original_notes = deal.get("notes", "") if deal else ""
        self.original_deal_id = deal.get("id", "") if deal else ""
        self.existing_deals = [dict(item) for item in (existing_deals or [])]
        self.notes_history_entries: list[dict[str, str]] = []
        self.title("ثبت معامله" if not deal else "ویرایش معامله")
        self.geometry("1280x920")
        self.transient(master)
        self.grab_set()
        bind_edit_shortcuts(self)

        wrapper = tb.Frame(self, padding=16)
        wrapper.pack(fill=BOTH, expand=True)

        self.inputs: dict[str, object] = {}
        combo_map = {
            "deal_type": [item["title"] for item in deal_types],
            "category": [item["title"] for item in categories],
            "pipeline": [item["title"] for item in pipelines],
            "status": DEAL_STATUSES,
        }
        form_grid = tb.Frame(wrapper)
        form_grid.pack(fill=X)
        form_grid.columnconfigure(0, weight=1)
        form_grid.columnconfigure(1, weight=1)
        fields = [
            ("title", "عنوان معامله", 0, 0),
            ("customer_name", "مشتری", 0, 1),
            ("product", "محصول", 1, 0),
            ("deal_type", "نوع معامله", 1, 1),
            ("category", "دسته‌بندی", 2, 0),
            ("pipeline", "کاریز", 2, 1),
            ("status", "وضعیت", 3, 0),
            ("sale_price", "قیمت فروش", 3, 1),
            ("operator_commission", "کمسیون اپراتور", 4, 1),
        ]
        for key, title, row, column in fields:
            cell = tb.Frame(form_grid, padding=6)
            cell.grid(row=row, column=column, sticky="ew")
            rtl_label(cell, title).pack(fill=X, pady=(0, 4))
            if key == "customer_name":
                widget = SearchableCombobox(cell, values=self.recent_customer_values(), justify="right")
            elif key in combo_map:
                widget = ttk.Combobox(cell, values=combo_map[key], justify="right", state="readonly")
            elif key == "product":
                widget = ttk.Combobox(cell, values=[], justify="right", state="readonly")
            else:
                widget = tb.Entry(cell, justify="right")
            widget.pack(fill=X, pady=(0, 10))
            self.inputs[key] = widget
        self.inputs["customer_name"].bind("<FocusIn>", lambda _event: self.inputs["customer_name"].set_values(self.recent_customer_values()))
        self.inputs["deal_type"].bind("<<ComboboxSelected>>", lambda _event: self.refresh_product_options())
        self.inputs["sale_price"].bind("<KeyRelease>", self.on_price_change)
        self.inputs["operator_commission"].bind("<KeyRelease>", self.on_price_change)

        bottom_grid = tb.Frame(wrapper)
        bottom_grid.pack(fill=BOTH, expand=True, pady=(6, 0))
        bottom_grid.columnconfigure(0, weight=1)
        bottom_grid.columnconfigure(1, weight=2)

        reminder_box = tb.Frame(bottom_grid, padding=6)
        reminder_box.grid(row=0, column=0, sticky="nsew")
        rtl_label(reminder_box, "تاریخ یادآور").pack(fill=X, pady=(0, 4))
        reminder_date = JalaliDateField(reminder_box, reminder_counter=self.reminder_counts_for_month)
        reminder_date.pack(fill=X, pady=(0, 8))
        self.inputs["reminder_date"] = reminder_date

        time_row = tb.Frame(reminder_box)
        time_row.pack(fill=X, pady=(0, 10))
        rtl_label(time_row, "ساعت", width=10).pack(side=RIGHT)
        reminder_hour = tb.Spinbox(time_row, from_=0, to=23, width=5, justify="center")
        reminder_hour.pack(side=RIGHT, padx=6)
        rtl_label(time_row, "دقیقه", width=10).pack(side=RIGHT)
        reminder_minute = tb.Spinbox(time_row, from_=0, to=59, width=5, justify="center")
        reminder_minute.pack(side=RIGHT, padx=6)
        self.inputs["reminder_hour"] = reminder_hour
        self.inputs["reminder_minute"] = reminder_minute

        messenger_box = tb.LabelFrame(reminder_box, text="ارتباط سریع")
        messenger_box.pack(fill=X, pady=(8, 0))
        messenger_body = tb.Frame(messenger_box, padding=8)
        messenger_body.pack(fill=X)
        tb.Button(messenger_body, text="Telegram", bootstyle="info-outline", command=self.open_telegram).pack(side=RIGHT, padx=4, pady=2)
        tb.Button(messenger_body, text="WhatsApp", bootstyle="success-outline", command=self.open_whatsapp).pack(side=RIGHT, padx=4, pady=2)
        tb.Button(messenger_body, text="Bale", bootstyle="primary-outline", command=self.open_bale).pack(side=RIGHT, padx=4, pady=2)
        tb.Button(messenger_body, text="Eitaa", bootstyle="warning-outline", command=self.open_eitaa).pack(side=RIGHT, padx=4, pady=2)
        tb.Button(messenger_body, text="Rubika", bootstyle="danger-outline", command=self.open_rubika).pack(side=RIGHT, padx=4, pady=2)

        action_box = tb.LabelFrame(reminder_box, text="ثبت معامله")
        action_box.pack(fill=X, pady=(10, 0))
        action_body = tb.Frame(action_box, padding=10)
        action_body.pack(fill=X)
        tb.Button(action_body, text="ذخیره", width=18, bootstyle="success", command=self.save).pack(fill=X, pady=(0, 8))
        tb.Button(action_body, text="انصراف", width=18, bootstyle="secondary", command=self.destroy).pack(fill=X)

        notes_box = tb.Frame(bottom_grid, padding=6)
        notes_box.grid(row=0, column=1, sticky="nsew")
        rtl_label(notes_box, "یادداشت").pack(fill=X, pady=(0, 4))
        history_box = tb.LabelFrame(notes_box, text="یادداشت‌ها")
        history_box.pack(fill=BOTH, expand=True, pady=(10, 0))
        history_body = tb.Frame(history_box, padding=8)
        history_body.pack(fill=BOTH, expand=True)
        history_toolbar = tb.Frame(history_body)
        history_toolbar.pack(fill=X, pady=(0, 8))
        tb.Button(history_toolbar, text="افزودن یادداشت", bootstyle="success", command=self.add_note).pack(side=RIGHT, padx=4)
        tb.Button(history_toolbar, text="ویرایش یادداشت", bootstyle="warning", command=self.edit_selected_note).pack(side=RIGHT, padx=4)
        tb.Button(history_toolbar, text="حذف یادداشت", bootstyle="danger", command=self.delete_selected_note).pack(side=RIGHT, padx=4)
        self.notes_history_tree = ttk.Treeview(history_body, columns=("time", "preview"), show="headings", height=6)
        self.notes_history_tree.heading("time", text="زمان ثبت")
        self.notes_history_tree.heading("preview", text="خلاصه یادداشت")
        self.notes_history_tree.column("time", width=180, anchor="e")
        self.notes_history_tree.column("preview", width=420, anchor="e")
        self.notes_history_tree.pack(fill=BOTH, expand=True)
        self.notes_history_tree.bind("<<TreeviewSelect>>", self.show_history_note_preview)
        self.notes_history_tree.bind("<Double-1>", lambda _event: self.edit_selected_note())
        preview_box = tb.LabelFrame(history_body, text="نمایش کامل یادداشت انتخاب‌شده")
        preview_box.pack(fill=BOTH, expand=True, pady=(8, 0))
        preview_body = tb.Frame(preview_box, padding=8)
        preview_body.pack(fill=BOTH, expand=True)
        preview_scroll = tb.Scrollbar(preview_body, orient="vertical")
        preview_scroll.pack(side=LEFT, fill=Y)
        self.note_history_preview = tk.Text(
            preview_body,
            height=8,
            wrap="word",
            font=("Tahoma", 10),
            padx=10,
            pady=8,
            state="disabled",
            yscrollcommand=preview_scroll.set,
        )
        self.note_history_preview.pack(side=RIGHT, fill=BOTH, expand=True)
        preview_scroll.configure(command=self.note_history_preview.yview)

        bind_primary_action(self, self.save)

        if deal:
            for key, widget in self.inputs.items():
                if key in {"reminder_date", "reminder_hour", "reminder_minute"}:
                    continue
                if hasattr(widget, "set"):
                    widget.set(deal.get(key, ""))
                else:
                    widget.insert(0, deal.get(key, ""))
            self.refresh_product_options(selected_product=deal.get("product", ""))
            self.notes_history_entries = list(reversed(parse_notes_history(deal.get("notes", ""))))
            self.refresh_notes_history()
            reminder_at = deal.get("reminder_at", "")
            when = parse_jalali_datetime(reminder_at)
            if when:
                jalali_when = jdatetime.datetime.fromgregorian(datetime=when)
                reminder_date.set(jalali_when.strftime("%Y/%m/%d"))
                reminder_hour.delete(0, END)
                reminder_hour.insert(0, f"{jalali_when.hour:02d}")
                reminder_minute.delete(0, END)
                reminder_minute.insert(0, f"{jalali_when.minute:02d}")
        else:
            self.inputs["status"].set(DEAL_STATUSES[0])
            if categories:
                self.inputs["category"].set(categories[0]["title"])
            if pipelines:
                self.inputs["pipeline"].set(pipelines[0]["title"])
            if deal_types:
                self.inputs["deal_type"].set(deal_types[0]["title"])
            self.refresh_product_options()
            self.inputs["customer_name"].set("")
            reminder_hour.insert(0, "09")
            reminder_minute.insert(0, "00")
            self.refresh_notes_history()

        self.wait_window(self)

    def save(self):
        title = self.inputs["title"].get().strip()
        customer_name = self.inputs["customer_name"].get().strip()
        if not title or not customer_name:
            Messagebox.show_error("عنوان معامله و مشتری الزامی است.", "خطا", parent=self)
            return
        try:
            hour = int(self.inputs["reminder_hour"].get())
            minute = int(self.inputs["reminder_minute"].get())
            reminder_at = f"{self.inputs['reminder_date'].get().strip()} {hour:02d}:{minute:02d}"
            if parse_jalali_datetime(reminder_at) is None:
                raise ValueError
        except ValueError:
            Messagebox.show_error("تاریخ یا ساعت یادآور معتبر نیست.", "خطا", parent=self)
            return

        notes_value = serialize_notes_history(list(reversed(self.notes_history_entries)))
        self.result = {
            "title": title,
            "customer_name": customer_name,
            "deal_type": self.inputs["deal_type"].get().strip(),
            "product": self.inputs["product"].get().strip(),
            "category": self.inputs["category"].get().strip(),
            "pipeline": self.inputs["pipeline"].get().strip(),
            "status": self.inputs["status"].get().strip(),
            "sale_price": self.inputs["sale_price"].get().strip(),
            "operator_commission": self.inputs["operator_commission"].get().strip(),
            "sales_expert": "",
            "notes": notes_value,
            "reminder_at": reminder_at,
        }
        self.destroy()

    def refresh_product_options(self, selected_product: str = ""):
        selected_type = self.inputs["deal_type"].get().strip()
        values = [item["title"] for item in self.products if item["deal_type"] == selected_type]
        self.inputs["product"].configure(values=values)
        if selected_product and selected_product in values:
            self.inputs["product"].set(selected_product)
        elif values:
            self.inputs["product"].set(values[0])
        else:
            self.inputs["product"].set("")

    def on_price_change(self, _event=None):
        for key in ("sale_price", "operator_commission"):
            entry = self.inputs.get(key)
            if not entry:
                continue
            current = entry.get()
            formatted = format_price_input(current)
            if current != formatted:
                entry.delete(0, END)
                entry.insert(0, formatted)

    def refresh_notes_history(self):
        for item in self.notes_history_tree.get_children():
            self.notes_history_tree.delete(item)
        for index, entry in enumerate(self.notes_history_entries):
            self.notes_history_tree.insert("", END, iid=str(index), values=(entry["time"], entry["preview"]))
        if self.notes_history_entries:
            first = self.notes_history_tree.get_children()[0]
            self.notes_history_tree.selection_set(first)
            self.show_history_note_preview()
        else:
            self.set_history_preview("")

    def set_history_preview(self, text: str):
        self.note_history_preview.configure(state="normal")
        self.note_history_preview.delete("1.0", END)
        payload = decode_rich_note(text)
        self.note_history_preview.tag_configure("rtl", justify="right", rmargin=14, lmargin1=14, lmargin2=14, spacing1=3, spacing3=4)
        self.note_history_preview.tag_configure("bold", font=("Tahoma", 10, "bold"))
        self.note_history_preview.tag_configure("italic", font=("Tahoma", 10, "italic"))
        self.note_history_preview.tag_configure("underline", underline=True)
        self.note_history_preview.tag_configure("heading", font=("Tahoma", 13, "bold"), spacing1=8, spacing3=6)
        self.note_history_preview.tag_configure("link", foreground="#0d6efd", underline=True)
        self.note_history_preview.tag_configure("timestamp", foreground="#198754", font=("Tahoma", 9, "bold"))
        self.note_history_preview.tag_configure("quote", background="#f8f9fa", foreground="#495057", lmargin1=28, lmargin2=28, rmargin=18)
        if payload:
            content = payload.get("text", "")
            self.note_history_preview.insert("1.0", content)
            for span in payload.get("spans", []):
                tag_name = span.get("tag")
                start = f"1.0+{int(span.get('start', 0))}c"
                end = f"1.0+{int(span.get('end', 0))}c"
                self.note_history_preview.tag_add(tag_name, start, end)
        else:
            self.note_history_preview.insert("1.0", text)
        self.note_history_preview.tag_add("rtl", "1.0", "end")
        self.note_history_preview.configure(state="disabled")

    def show_history_note_preview(self, _event=None):
        selected = self.notes_history_tree.selection()
        if not selected:
            self.set_history_preview("")
            return
        entry = self.notes_history_entries[int(selected[0])]
        body = entry["text"].strip() or "-"
        preview_text = body
        self.set_history_preview(preview_text)

    def selected_note_index(self) -> int | None:
        selected = self.notes_history_tree.selection()
        if not selected:
            return None
        return int(selected[0])

    def add_note(self):
        dialog = NoteEditorDialog(self)
        if not dialog.result:
            return
        self.notes_history_entries.insert(0, dialog.result)
        self.refresh_notes_history()

    def edit_selected_note(self):
        note_index = self.selected_note_index()
        if note_index is None:
            Messagebox.show_warning("ابتدا یک یادداشت را انتخاب کنید.", "یادداشت", parent=self)
            return
        dialog = NoteEditorDialog(self, self.notes_history_entries[note_index])
        if not dialog.result:
            return
        self.notes_history_entries[note_index] = dialog.result
        self.refresh_notes_history()
        self.notes_history_tree.selection_set(str(note_index))
        self.show_history_note_preview()

    def delete_selected_note(self):
        note_index = self.selected_note_index()
        if note_index is None:
            Messagebox.show_warning("ابتدا یک یادداشت را انتخاب کنید.", "یادداشت", parent=self)
            return
        if Messagebox.yesno("یادداشت انتخاب‌شده حذف شود؟", "یادداشت", parent=self) != "Yes":
            return
        del self.notes_history_entries[note_index]
        self.refresh_notes_history()

    def reminder_counts_for_month(self, year: int, month: int) -> dict[int, int]:
        counts: dict[int, int] = defaultdict(int)
        for item in self.existing_deals:
            if item.get("id") == self.original_deal_id:
                continue
            due = parse_jalali_datetime(item.get("reminder_at", ""))
            if due is None:
                continue
            jalali_due = jdatetime.datetime.fromgregorian(datetime=due)
            if jalali_due.year == year and jalali_due.month == month:
                counts[jalali_due.day] += 1
        return dict(counts)

    def recent_customer_values(self) -> list[str]:
        ordered = sorted(self.customers, key=lambda item: item.get("updated_at", item.get("created_at", "")), reverse=True)
        names = []
        seen = set()
        for item in ordered:
            name = (item.get("name") or "").strip()
            if not name or name in seen:
                continue
            seen.add(name)
            names.append(name)
        return names

    def current_customer_mobile(self) -> str:
        customer = self.customer_lookup.get(self.inputs["customer_name"].get().strip())
        return customer.get("mobile", "") if customer else ""

    def open_telegram(self):
        mobile = "".join(ch for ch in self.current_customer_mobile() if ch.isdigit() or ch == "+")
        if mobile:
            webbrowser.open(f"tg://resolve?phone={mobile}")

    def open_whatsapp(self):
        mobile = "".join(ch for ch in self.current_customer_mobile() if ch.isdigit())
        if mobile:
            webbrowser.open(f"https://wa.me/{mobile}")

    def open_bale(self):
        mobile = "".join(ch for ch in self.current_customer_mobile() if ch.isdigit())
        if mobile:
            webbrowser.open(f"https://ble.ir/{mobile}")
        else:
            webbrowser.open("https://web.bale.ai/")

    def open_eitaa(self):
        mobile = "".join(ch for ch in self.current_customer_mobile() if ch.isdigit())
        eitaa_path = Path.home() / "AppData" / "Roaming" / "Eitaa Desktop" / "Eitaa.exe"
        if eitaa_path.exists():
            try:
                subprocess.Popen([str(eitaa_path)])
                return
            except Exception:
                pass
        if mobile:
            try:
                os.startfile(f"eitaa://chat/{mobile}")
                return
            except Exception:
                pass
            webbrowser.open(f"eitaa://chat/{mobile}")
        else:
            try:
                os.startfile("eitaa://")
            except Exception:
                pass

    def open_rubika(self):
        mobile = "".join(ch for ch in self.current_customer_mobile() if ch.isdigit())
        if mobile:
            webbrowser.open(f"https://web.rubika.ir/#{mobile}")
        else:
            webbrowser.open("https://web.rubika.ir/")


class DashboardCard(tb.Frame):
    def __init__(self, master, title: str, value: str, accent: str):
        super().__init__(master, bootstyle=accent, padding=14)
        tb.Label(self, text=title, font=("Tahoma", 10), anchor="e").pack(fill=X)
        self.value_label = tb.Label(self, text=value, font=("Tahoma", 20, "bold"), anchor="e")
        self.value_label.pack(fill=X, pady=(8, 0))

    def set_value(self, value: str):
        self.value_label.configure(text=value)


class TooltipWindow:
    def __init__(self, master):
        self.master = master
        self.window: tk.Toplevel | None = None
        self.label: tk.Label | None = None

    def show(self, text: str, x: int, y: int):
        if not text:
            self.hide()
            return
        if self.window is None or not self.window.winfo_exists():
            self.window = tk.Toplevel(self.master)
            self.window.withdraw()
            self.window.overrideredirect(True)
            self.window.attributes("-topmost", True)
            self.label = tk.Label(
                self.window,
                justify="right",
                anchor="e",
                bg="#fffdf5",
                fg="#111111",
                relief="solid",
                bd=1,
                padx=10,
                pady=8,
                font=("Tahoma", 9),
                wraplength=420,
            )
            self.label.pack(fill=BOTH, expand=True)
        self.label.configure(text=text)
        self.window.geometry(f"+{x}+{y}")
        self.window.deiconify()

    def hide(self):
        if self.window is not None and self.window.winfo_exists():
            self.window.withdraw()


class BarChart(tb.Frame):
    def __init__(self, master, title: str):
        super().__init__(master, padding=10)
        rtl_label(self, title, font=("Tahoma", 11, "bold")).pack(fill=X, pady=(0, 8))
        self.canvas = tk.Canvas(self, height=240, highlightthickness=0, bg="#ffffff")
        self.canvas.pack(fill=BOTH, expand=True)

    def render(self, data: list[tuple[str, float]]):
        self.canvas.delete("all")
        if not data:
            self.canvas.create_text(250, 120, text="داده‌ای برای نمایش وجود ندارد", fill="#6c757d", font=("Tahoma", 11))
            return
        width = max(self.canvas.winfo_width(), 500)
        height = max(self.canvas.winfo_height(), 240)
        max_value = max(value for _label, value in data) or 1
        bar_width = max(38, int((width - 80) / max(len(data), 1)) - 14)
        x = width - 60
        colors = ["#0d6efd", "#20c997", "#fd7e14", "#d63384", "#6f42c1", "#198754", "#dc3545"]
        for index, (label, value) in enumerate(data):
            bar_height = int((height - 80) * (value / max_value))
            left = x - bar_width
            top = height - 46 - bar_height
            self.canvas.create_rectangle(left, top, x, height - 46, fill=colors[index % len(colors)], outline="")
            self.canvas.create_text((left + x) / 2, top - 12, text=f"{value:,.0f}", font=("Tahoma", 9))
            self.canvas.create_text((left + x) / 2, height - 22, text=label, font=("Tahoma", 9))
            x -= bar_width + 18


class LineChart(tb.Frame):
    def __init__(self, master, title: str):
        super().__init__(master, padding=10)
        rtl_label(self, title, font=("Tahoma", 11, "bold")).pack(fill=X, pady=(0, 8))
        self.canvas = tk.Canvas(self, height=220, highlightthickness=0, bg="#ffffff")
        self.canvas.pack(fill=BOTH, expand=True)

    def render(self, data: list[tuple[str, float]]):
        self.canvas.delete("all")
        if not data:
            self.canvas.create_text(250, 110, text="داده‌ای برای نمایش وجود ندارد", fill="#6c757d", font=("Tahoma", 11))
            return
        width = max(self.canvas.winfo_width(), 500)
        height = max(self.canvas.winfo_height(), 220)
        margin = 40
        max_value = max(value for _label, value in data) or 1
        step = (width - 2 * margin) / max(1, len(data) - 1)
        points = []
        for index, (label, value) in enumerate(data):
            x = margin + index * step
            y = height - margin - ((height - 2 * margin) * (value / max_value))
            points.extend([x, y])
            self.canvas.create_oval(x - 4, y - 4, x + 4, y + 4, fill="#0d6efd", outline="")
            self.canvas.create_text(x, height - 18, text=label, font=("Tahoma", 8))
        if len(points) >= 4:
            self.canvas.create_line(*points, fill="#0d6efd", width=3, smooth=True)


class ColumnOrderDialog(tb.Toplevel):
    def __init__(self, master, tree: ttk.Treeview, title: str):
        super().__init__(master)
        self.tree = tree
        self.result = None
        self.title(title)
        self.geometry("320x360")
        self.transient(master)
        self.grab_set()
        wrapper = tb.Frame(self, padding=14)
        wrapper.pack(fill=BOTH, expand=True)
        self.listbox = tk.Listbox(wrapper)
        self.listbox.pack(fill=BOTH, expand=True)
        self.column_map = {}
        displaycolumns = list(tree["displaycolumns"])
        if displaycolumns == ["#all"] or displaycolumns == ("#all",):
            displaycolumns = list(tree["columns"])
        for col in displaycolumns:
            label = tree.heading(col)["text"] or col
            self.column_map[label] = col
            self.listbox.insert(END, label)
        actions = tb.Frame(wrapper)
        actions.pack(fill=X, pady=(10, 0))
        tb.Button(actions, text="بالا", bootstyle="info-outline", command=lambda: self.move(-1)).pack(side=RIGHT, padx=4)
        tb.Button(actions, text="پایین", bootstyle="info-outline", command=lambda: self.move(1)).pack(side=RIGHT, padx=4)
        tb.Button(actions, text="اعمال", bootstyle="success", command=self.apply).pack(side=RIGHT, padx=4)
        tb.Button(actions, text="بستن", bootstyle="secondary", command=self.destroy).pack(side=LEFT, padx=4)
        self.wait_window(self)

    def move(self, offset: int):
        selected = self.listbox.curselection()
        if not selected:
            return
        index = selected[0]
        target = index + offset
        if target < 0 or target >= self.listbox.size():
            return
        value = self.listbox.get(index)
        self.listbox.delete(index)
        self.listbox.insert(target, value)
        self.listbox.selection_set(target)

    def apply(self):
        self.tree["displaycolumns"] = [self.column_map[label] for label in self.listbox.get(0, END)]
        self.destroy()


class InAppNotification(tb.Toplevel):
    def __init__(self, master, title: str, message: str, click_callback: Callable[[], None]):
        super().__init__(master)
        self.click_callback = click_callback
        self.overrideredirect(True)
        self.attributes("-topmost", True)
        self.configure(bg="#ffffff")
        width, height = 360, 120
        x = self.winfo_screenwidth() - width - 24
        y = self.winfo_screenheight() - height - 70
        self.geometry(f"{width}x{height}+{x}+{y}")
        wrapper = tb.Frame(self, padding=12, bootstyle="light")
        wrapper.pack(fill=BOTH, expand=True)
        top = tb.Frame(wrapper)
        top.pack(fill=X)
        tb.Button(top, text="Ã—", width=3, bootstyle="danger-outline", command=self.destroy).pack(side=LEFT)
        tb.Label(top, text=title, font=("Tahoma", 10, "bold"), anchor="e").pack(side=RIGHT, fill=X, expand=True)
        body = tb.Label(wrapper, text=message, justify="right", anchor="e", wraplength=300)
        body.pack(fill=BOTH, expand=True, pady=(8, 0))
        for widget in (wrapper, top, body):
            widget.bind("<Button-1>", self.on_click)

    def on_click(self, _event=None):
        self.destroy()
        self.click_callback()


class CRMApp(tb.Toplevel):
    def __init__(self, master, profile: ProfileContext):
        super().__init__(master)
        self.profile = profile
        self.store = CrmStore(profile)
        self.reminder_poll_ms = 5000
        self.app_started_at = datetime.now()
        self.settings_unlocked = False
        self.tray_icon = None
        self.tray_thread = None
        self.in_app_notifications: list[InAppNotification] = []
        self.deal_notes_tooltip = TooltipWindow(self)
        self.window_icon = None
        self.customer_page = 1
        self.deal_page = 1
        self.list_page_size = 12
        self.settings_path = app_settings_path_for_profile(profile)
        self.app_settings = load_app_settings(self.settings_path)
        self.startup_enabled_var = tk.BooleanVar(value=self.app_settings.get("launch_on_startup", True))
        self.title(f"{APP_TITLE} - {profile.title}")
        self.geometry("1480x1000")
        self.minsize(1200, 980)
        try:
            self.state("zoomed")
        except tk.TclError:
            self.attributes("-fullscreen", True)
        self.option_add("*Font", "Tahoma 10")
        self.style.configure("Treeview", rowheight=28)
        self.style.configure("Treeview", font=("Tahoma", 10))
        self.style.configure("TNotebook", tabposition="ne")
        self.style.configure("Treeview.Heading", anchor="e")
        self.style.configure("Treeview.Heading", font=("Tahoma", 10, "bold"))
        self.style.configure("TLabel", font=("Tahoma", 10))
        self.style.configure("TLabelframe.Label", font=("Tahoma", 10, "bold"))
        self.style.configure("TButton", font=("Tahoma", 10))
        self.style.configure("TCheckbutton", font=("Tahoma", 10))
        self.style.configure("TCombobox", font=("Tahoma", 10))
        self.configure(bg="#f7f7fb")
        self.style.configure("Header.TLabel", font=("Tahoma", 18, "bold"), anchor="e")
        self.style.configure("Section.TLabel", font=("Tahoma", 12, "bold"), anchor="e")
        self.style.configure("Sidebar.TButton", font=("Tahoma", 10, "bold"), foreground="#111111", background="#ffffff")
        self.style.configure("SidebarTitle.TLabel", font=("Tahoma", 12, "bold"), foreground="#111111", anchor="e")
        try:
            Registry(APP_NOTIFY_ID, sys.executable, os.path.abspath(sys.argv[0]), force_override=True)
        except Exception:
            pass
        self.apply_window_icon()
        bind_edit_shortcuts(self)
        set_windows_startup(self.startup_enabled_var.get())
        self.build_layout()
        self.refresh_everything()
        self.protocol("WM_DELETE_WINDOW", self.close_app)
        self.notebook.bind("<<NotebookTabChanged>>", self.on_tab_changed)
        self.after(self.reminder_poll_ms, self.check_reminders)

    def close_app(self):
        for popup in list(self.in_app_notifications):
            try:
                popup.destroy()
            except Exception:
                pass
        if self.tray_icon:
            try:
                self.tray_icon.stop()
            except Exception:
                pass
        try:
            self.master.destroy()
        except Exception:
            self.destroy()

    def save_app_settings(self):
        self.app_settings["launch_on_startup"] = bool(self.startup_enabled_var.get())
        save_app_settings(self.settings_path, self.app_settings)

    def toggle_startup_setting(self):
        enabled = bool(self.startup_enabled_var.get())
        if not set_windows_startup(enabled):
            Messagebox.show_error("تنظیم استارت‌آپ ویندوز انجام نشد.", "تنظیمات", parent=self)
            self.startup_enabled_var.set(not enabled)
            return
        self.save_app_settings()

    def current_page_size(self, tree: ttk.Treeview) -> int:
        try:
            rowheight = int(self.style.lookup("Treeview", "rowheight") or 28)
            visible_height = max(0, tree.winfo_height() - 34)
            return max(1, visible_height // max(rowheight, 1))
        except Exception:
            return self.list_page_size

    def create_tray_image(self):
        icon_path = get_app_icon_path()
        if icon_path:
            try:
                return Image.open(icon_path).convert("RGBA")
            except Exception:
                pass
        image = Image.new("RGB", (64, 64), "#1d3557")
        draw = ImageDraw.Draw(image)
        draw.rounded_rectangle((8, 8, 56, 56), radius=10, fill="#2a9d8f")
        draw.text((18, 20), "OC", fill="white")
        return image

    def apply_window_icon(self):
        icon_path = get_app_icon_path()
        if not icon_path:
            return
        try:
            self.window_icon = tk.PhotoImage(file=str(icon_path))
            self.iconphoto(True, self.window_icon)
            if hasattr(self.master, "iconphoto"):
                self.master.iconphoto(True, self.window_icon)
        except Exception:
            pass

    def hide_to_tray(self):
        if self.tray_icon:
            return
        self.withdraw()
        menu = pystray.Menu(
            pystray.MenuItem("باز کردن", lambda icon, item: self.after(0, self.restore_from_tray), default=True),
            pystray.MenuItem("داشبورد", lambda icon, item: self.after(0, lambda: self.restore_from_tray(self.dashboard_tab))),
            pystray.MenuItem("مشتریان", lambda icon, item: self.after(0, lambda: self.restore_from_tray(self.customers_tab))),
            pystray.MenuItem("معاملات", lambda icon, item: self.after(0, lambda: self.restore_from_tray(self.deals_tab))),
            pystray.MenuItem("فروش و درآمد", lambda icon, item: self.after(0, lambda: self.restore_from_tray(self.sales_tab))),
            pystray.MenuItem("خروج", lambda icon, item: self.after(0, self.close_app)),
        )
        self.tray_icon = pystray.Icon(APP_NOTIFY_ID, self.create_tray_image(), APP_TITLE, menu)
        self.tray_icon.run_detached()

    def restore_from_tray(self, tab=None):
        self.deiconify()
        self.lift()
        self.focus_force()
        if tab is not None:
            self.notebook.select(tab)
        if self.tray_icon:
            try:
                self.tray_icon.stop()
            except Exception:
                pass
            self.tray_icon = None

    def app_is_minimized(self) -> bool:
        try:
            return self.state() == "iconic" or not self.winfo_viewable() or self.tray_icon is not None
        except Exception:
            return False

    def clear_dead_notifications(self):
        self.in_app_notifications = [popup for popup in self.in_app_notifications if popup.winfo_exists()]

    def show_in_app_notification(self, title: str, message: str, callback: Callable[[], None]):
        self.clear_dead_notifications()
        popup = InAppNotification(self.master, title, message, callback)
        offset = len(self.in_app_notifications) * 128
        width, height = 360, 120
        x = popup.winfo_screenwidth() - width - 24
        y = popup.winfo_screenheight() - height - 70 - offset
        popup.geometry(f"{width}x{height}+{x}+{y}")
        self.in_app_notifications.append(popup)
        if winsound:
            try:
                winsound.MessageBeep()
            except RuntimeError:
                pass

    def on_tab_changed(self, _event=None):
        current = self.notebook.select()
        if current == str(self.settings_tab):
            if not self.settings_unlocked:
                auth = SimpleLoginDialog(self, "ورود به تنظیمات", default_username="admin")
                if not auth.result or auth.result["username"] != "admin" or auth.result["password"] != "admin":
                    Messagebox.show_error("دسترسی به تنظیمات مجاز نیست.", "تنظیمات", parent=self)
                    self.notebook.select(self.dashboard_tab)
                else:
                    self.settings_unlocked = True
        else:
            self.settings_unlocked = False

    def build_layout(self):
        wrapper = tb.Frame(self, padding=16)
        wrapper.pack(fill=BOTH, expand=True)

        header = tb.Frame(wrapper, padding=(16, 14), bootstyle="light")
        header.pack(fill=X, pady=(0, 16))
        header_right = tb.Frame(header)
        header_right.pack(side=RIGHT)
        tb.Button(header_right, text="ارسال به Tray", bootstyle="secondary-outline", command=self.hide_to_tray).pack(side=RIGHT, padx=(0, 8))
        tb.Label(header_right, text=APP_TITLE, style="Header.TLabel").pack(side=RIGHT, padx=(0, 12))
        rtl_label(header_right, f"پروفایل فعال: {self.profile.title}", bootstyle="secondary").pack(side=RIGHT)
        self.notification_badge = tb.Label(header, text=rtl_text("اعلان‌های نخوانده: 0"), bootstyle="danger")
        self.notification_badge.pack(side=LEFT)

        content = tb.Frame(wrapper)
        content.pack(fill=BOTH, expand=True)
        sidebar = tb.Frame(content, padding=12, bootstyle="dark")
        sidebar.pack(side=RIGHT, fill=Y, padx=(12, 0))
        tb.Label(sidebar, text="بخش‌ها", style="SidebarTitle.TLabel").pack(fill=X, pady=(0, 8))

        self.notebook = ttk.Notebook(content)
        self.notebook.pack(side=LEFT, fill=BOTH, expand=True)
        self.dashboard_tab = tb.Frame(self.notebook, padding=14)
        self.customers_tab = tb.Frame(self.notebook, padding=14)
        self.deals_tab = tb.Frame(self.notebook, padding=14)
        self.sales_tab = tb.Frame(self.notebook, padding=14)
        self.reports_tab = tb.Frame(self.notebook, padding=14)
        self.notifications_tab = tb.Frame(self.notebook, padding=14)
        self.io_tab = tb.Frame(self.notebook, padding=14)
        self.settings_tab = tb.Frame(self.notebook, padding=14)
        self.guide_tab = tb.Frame(self.notebook, padding=14)
        self.about_tab = tb.Frame(self.notebook, padding=14)
        for tab, title in [
            (self.dashboard_tab, "داشبورد"),
            (self.customers_tab, "مشتریان"),
            (self.deals_tab, "معاملات"),
            (self.sales_tab, "فروش و درآمد"),
            (self.reports_tab, "گزارشات"),
            (self.notifications_tab, "اعلان‌ها"),
            (self.io_tab, "ورودی/خروجی"),
            (self.settings_tab, "تنظیمات"),
            (self.guide_tab, "راهنما"),
            (self.about_tab, "درباره ما"),
        ]:
            self.notebook.add(tab, text=title)
            tb.Button(sidebar, text=title, width=18, style="Sidebar.TButton", bootstyle="light", command=lambda t=tab: self.notebook.select(t)).pack(fill=X, pady=4)

        self.build_dashboard()
        self.build_customers()
        self.build_deals()
        self.build_sales()
        self.build_reports()
        self.build_notifications()
        self.build_io()
        self.build_settings()
        self.build_guide()
        self.build_about()

    def build_dashboard(self):
        rtl_label(self.dashboard_tab, "داشبورد حرفه‌ای فروش", style="Section.TLabel").pack(fill=X)
        toolbar = tb.Frame(self.dashboard_tab)
        toolbar.pack(fill=X, pady=(14, 10))
        tb.Button(toolbar, text="ثبت معامله", bootstyle="success", command=self.add_deal).pack(side=RIGHT, padx=4)
        tb.Button(toolbar, text="ویرایش معامله", bootstyle="warning", command=lambda: self.edit_selected_deal_from(self.dashboard_deal_tree)).pack(side=RIGHT, padx=4)
        tb.Button(toolbar, text="حذف معامله", bootstyle="danger", command=self.delete_dashboard_deal).pack(side=RIGHT, padx=4)
        tb.Button(toolbar, text="چیدمان ستون‌ها", bootstyle="secondary-outline", command=lambda: ColumnOrderDialog(self, self.dashboard_deal_tree, "چیدمان ستون‌های داشبورد")).pack(side=RIGHT, padx=4)
        self.dashboard_search = tb.Entry(toolbar, justify="right")
        self.dashboard_search.pack(side=LEFT, padx=6)
        self.dashboard_search.bind("<KeyRelease>", lambda _event: self.refresh_dashboard())

        filter_row = tb.Frame(self.dashboard_tab)
        filter_row.pack(fill=X, pady=(0, 14))
        self.dashboard_sort = ttk.Combobox(filter_row, state="readonly", justify="right", width=16, values=["بدون سورت", "یادآور", "وضعیت", "دسته‌بندی", "محصول", "نوع معامله"])
        self.dashboard_sort.pack(side=LEFT, padx=6)
        self.dashboard_sort.set("بدون سورت")
        self.dashboard_sort.bind("<<ComboboxSelected>>", lambda _event: self.refresh_dashboard())
        self.dashboard_status_filter = ttk.Combobox(filter_row, state="readonly", justify="right", width=13)
        self.dashboard_status_filter.pack(side=LEFT, padx=6)
        self.dashboard_status_filter.bind("<<ComboboxSelected>>", lambda _event: self.refresh_dashboard())
        self.dashboard_category_filter = ttk.Combobox(filter_row, state="readonly", justify="right", width=13)
        self.dashboard_category_filter.pack(side=LEFT, padx=6)
        self.dashboard_category_filter.bind("<<ComboboxSelected>>", lambda _event: self.refresh_dashboard())
        self.dashboard_product_filter = ttk.Combobox(filter_row, state="readonly", justify="right", width=13)
        self.dashboard_product_filter.pack(side=LEFT, padx=6)
        self.dashboard_product_filter.bind("<<ComboboxSelected>>", lambda _event: self.refresh_dashboard())
        self.dashboard_type_filter = ttk.Combobox(filter_row, state="readonly", justify="right", width=13)
        self.dashboard_type_filter.pack(side=LEFT, padx=6)
        self.dashboard_type_filter.bind("<<ComboboxSelected>>", lambda _event: self.refresh_dashboard())
        filter_row.pack_propagate(False)

        cards = tb.Frame(self.dashboard_tab)
        cards.pack(fill=X, pady=(0, 14))
        self.card_customers = DashboardCard(cards, "تعداد مشتریان", "0", "primary")
        self.card_deals = DashboardCard(cards, "تعداد معاملات", "0", "info")
        self.card_success = DashboardCard(cards, "در دست بررسی", "0", "success")
        self.card_revenue = DashboardCard(cards, "کمیسیون ماهیانه", "0", "warning")
        for card in [self.card_customers, self.card_deals, self.card_success, self.card_revenue]:
            card.pack(side=RIGHT, fill=X, expand=True, padx=5)

        deals_box = tb.LabelFrame(self.dashboard_tab, text="معاملات قابل نمایش در داشبورد")
        deals_box.pack(fill=BOTH, expand=True)
        deals_body = tb.Frame(deals_box, padding=10)
        deals_body.pack(fill=BOTH, expand=True)
        self.dashboard_deal_tree = ttk.Treeview(
            deals_body,
            columns=("title", "customer", "status", "category", "created", "reminder"),
            show="headings",
            height=8,
        )
        for key, title, width in [
            ("title", "عنوان", 220),
            ("customer", "مشتری", 180),
            ("status", "وضعیت", 120),
            ("category", "دسته‌بندی", 140),
            ("created", "تاریخ ایجاد", 120),
            ("reminder", "یادآور", 170),
        ]:
            self.dashboard_deal_tree.heading(key, text=title)
            self.dashboard_deal_tree.column(key, width=width, anchor="e")
        self.dashboard_deal_tree["displaycolumns"] = tuple(reversed(self.dashboard_deal_tree["columns"]))
        self.dashboard_deal_tree.pack(fill=BOTH, expand=True)
        self.dashboard_deal_tree.bind("<Double-1>", lambda _event: self.edit_selected_deal_from(self.dashboard_deal_tree))

    def build_customers(self):
        toolbar = tb.Frame(self.customers_tab)
        toolbar.pack(fill=X, pady=(0, 10))
        tb.Button(toolbar, text="ثبت مشتری", bootstyle="success", command=self.add_customer).pack(side=RIGHT, padx=4)
        tb.Button(toolbar, text="ویرایش مشتری", bootstyle="warning", command=self.edit_customer).pack(side=RIGHT, padx=4)
        tb.Button(toolbar, text="حذف مشتری", bootstyle="danger", command=self.delete_customer).pack(side=RIGHT, padx=4)
        tb.Button(toolbar, text="چیدمان ستون‌ها", bootstyle="secondary-outline", command=lambda: ColumnOrderDialog(self, self.customer_tree, "چیدمان ستون‌های مشتریان")).pack(side=RIGHT, padx=4)
        self.customer_search = tb.Entry(toolbar, justify="right")
        self.customer_search.pack(side=LEFT, padx=6)
        self.customer_search.bind("<KeyRelease>", lambda _event: self.reset_customer_page())
        self.customer_sort = ttk.Combobox(toolbar, state="readonly", justify="right", width=16, values=["بدون سورت", "نام", "موبایل", "تاریخ تولد", "سطح درآمد"])
        self.customer_sort.pack(side=LEFT, padx=6)
        self.customer_sort.set("بدون سورت")
        self.customer_sort.bind("<<ComboboxSelected>>", lambda _event: self.reset_customer_page())
        self.customer_tree = ttk.Treeview(self.customers_tab, columns=("name", "mobile", "birthdate", "income"), show="headings")
        for key, title, width in [("name", "نام", 220), ("mobile", "موبایل", 140), ("birthdate", "تاریخ تولد", 130), ("income", "سطح درآمد", 140)]:
            self.customer_tree.heading(key, text=title)
            self.customer_tree.column(key, width=width, anchor="e")
        self.customer_tree.pack(fill=BOTH, expand=True)
        self.customer_tree["displaycolumns"] = tuple(reversed(self.customer_tree["columns"]))
        self.customer_tree.bind("<Double-1>", lambda _event: self.edit_customer())
        self.customer_tree.bind("<Configure>", lambda _event: self.refresh_customer_tree())
        customer_pager = tb.Frame(self.customers_tab)
        customer_pager.pack(fill=X, pady=(8, 0))
        tb.Button(customer_pager, text="بعدی", bootstyle="secondary-outline", command=lambda: self.change_customer_page(1)).pack(side=LEFT, padx=4)
        tb.Button(customer_pager, text="قبلی", bootstyle="secondary-outline", command=lambda: self.change_customer_page(-1)).pack(side=LEFT, padx=4)
        tb.Button(customer_pager, text="تایید", bootstyle="info-outline", command=self.go_customer_page).pack(side=LEFT, padx=4)
        self.customer_page_entry = tb.Entry(customer_pager, width=6, justify="center")
        self.customer_page_entry.pack(side=LEFT, padx=4)
        self.customer_page_label = rtl_label(customer_pager, "صفحه 1 از 1")
        self.customer_page_label.pack(side=RIGHT, padx=6)
        self.customer_total_label = rtl_label(customer_pager, "تعداد کل: 0")
        self.customer_total_label.pack(side=RIGHT, padx=6)

    def build_deals(self):
        toolbar = tb.Frame(self.deals_tab)
        toolbar.pack(fill=X, pady=(0, 10))
        tb.Button(toolbar, text="ثبت معامله", bootstyle="success", command=self.add_deal).pack(side=RIGHT, padx=4)
        tb.Button(toolbar, text="ویرایش معامله", bootstyle="warning", command=self.edit_deal).pack(side=RIGHT, padx=4)
        tb.Button(toolbar, text="حذف معامله", bootstyle="danger", command=self.delete_deal).pack(side=RIGHT, padx=4)
        tb.Button(toolbar, text="چیدمان ستون‌ها", bootstyle="secondary-outline", command=lambda: ColumnOrderDialog(self, self.deal_tree, "چیدمان ستون‌های معاملات")).pack(side=RIGHT, padx=4)
        self.deal_search = tb.Entry(toolbar, justify="right")
        self.deal_search.pack(side=LEFT, padx=6)
        self.deal_search.bind("<KeyRelease>", lambda _event: self.reset_deal_page())
        self.deal_sort = ttk.Combobox(toolbar, state="readonly", justify="right", width=16, values=["بدون سورت", "یادآور", "وضعیت", "دسته‌بندی", "محصول", "نوع معامله"])
        self.deal_sort.pack(side=LEFT, padx=6)
        self.deal_sort.set("بدون سورت")
        self.deal_sort.bind("<<ComboboxSelected>>", lambda _event: self.reset_deal_page())
        self.deal_status_filter = ttk.Combobox(toolbar, state="readonly", justify="right", width=13)
        self.deal_status_filter.pack(side=LEFT, padx=6)
        self.deal_status_filter.bind("<<ComboboxSelected>>", lambda _event: self.reset_deal_page())
        self.deal_category_filter = ttk.Combobox(toolbar, state="readonly", justify="right", width=13)
        self.deal_category_filter.pack(side=LEFT, padx=6)
        self.deal_category_filter.bind("<<ComboboxSelected>>", lambda _event: self.reset_deal_page())
        self.deal_product_filter = ttk.Combobox(toolbar, state="readonly", justify="right", width=13)
        self.deal_product_filter.pack(side=LEFT, padx=6)
        self.deal_product_filter.bind("<<ComboboxSelected>>", lambda _event: self.reset_deal_page())
        self.deal_type_filter = ttk.Combobox(toolbar, state="readonly", justify="right", width=13)
        self.deal_type_filter.pack(side=LEFT, padx=6)
        self.deal_type_filter.bind("<<ComboboxSelected>>", lambda _event: self.reset_deal_page())
        self.deal_reminder_filter = JalaliDateField(toolbar, allow_empty=True, on_change=self.reset_deal_page)
        self.deal_reminder_filter.pack(side=LEFT, padx=6)
        self.deal_tree = ttk.Treeview(
            self.deals_tab,
            columns=("title", "customer", "type", "product", "category", "pipeline", "status", "price", "created", "reminder", "notes"),
            show="headings",
        )
        for key, title, width in [
            ("title", "عنوان", 170),
            ("customer", "مشتری", 140),
            ("type", "نوع", 90),
            ("product", "محصول", 120),
            ("category", "دسته‌بندی", 140),
            ("pipeline", "کاریز", 120),
            ("status", "وضعیت", 110),
            ("price", "قیمت فروش", 110),
            ("created", "تاریخ ایجاد", 120),
            ("reminder", "یادآور", 150),
        ]:
            self.deal_tree.heading(key, text=f"| {title} |")
            self.deal_tree.column(key, width=width, anchor="e")
        self.deal_tree.heading("notes", text="| یادداشت |")
        self.deal_tree.column("notes", width=240, anchor="e")
        self.deal_tree["displaycolumns"] = tuple(reversed(self.deal_tree["columns"]))
        self.deal_tree.pack(fill=BOTH, expand=True)
        self.deal_tree.bind("<Double-1>", lambda _event: self.edit_deal())
        self.deal_tree.bind("<Configure>", lambda _event: self.refresh_deal_tree())
        self.deal_tree.bind("<Motion>", self.on_deal_tree_motion)
        self.deal_tree.bind("<Leave>", lambda _event: self.deal_notes_tooltip.hide())
        deal_pager = tb.Frame(self.deals_tab)
        deal_pager.pack(fill=X, pady=(8, 0))
        tb.Button(deal_pager, text="بعدی", bootstyle="secondary-outline", command=lambda: self.change_deal_page(1)).pack(side=LEFT, padx=4)
        tb.Button(deal_pager, text="قبلی", bootstyle="secondary-outline", command=lambda: self.change_deal_page(-1)).pack(side=LEFT, padx=4)
        tb.Button(deal_pager, text="تایید", bootstyle="info-outline", command=self.go_deal_page).pack(side=LEFT, padx=4)
        self.deal_page_entry = tb.Entry(deal_pager, width=6, justify="center")
        self.deal_page_entry.pack(side=LEFT, padx=4)
        self.deal_page_label = rtl_label(deal_pager, "صفحه 1 از 1")
        self.deal_page_label.pack(side=RIGHT, padx=6)
        self.deal_total_label = rtl_label(deal_pager, "تعداد کل: 0")
        self.deal_total_label.pack(side=RIGHT, padx=6)

    def build_sales(self):
        self.sales_summary = rtl_label(self.sales_tab, "جمع فروش موفق: 0", style="Section.TLabel")
        self.sales_summary.pack(fill=X)
        cards = tb.Frame(self.sales_tab)
        cards.pack(fill=X, pady=(10, 12))
        self.sales_best_day = DashboardCard(cards, "پرفروش‌ترین روز", "-", "primary")
        self.sales_best_month = DashboardCard(cards, "پرفروش‌ترین ماه", "-", "info")
        self.sales_success_rate = DashboardCard(cards, "نرخ موفقیت", "0%", "success")
        self.sales_status_mix = DashboardCard(cards, "وضعیت‌ها", "-", "warning")
        for card in [self.sales_best_day, self.sales_best_month, self.sales_success_rate, self.sales_status_mix]:
            card.pack(side=RIGHT, fill=X, expand=True, padx=5)
        filters = tb.Frame(self.sales_tab)
        filters.pack(fill=X, pady=(10, 12))
        rtl_label(filters, "فیلتر نوع معامله").pack(side=RIGHT)
        self.sales_type_filter = ttk.Combobox(filters, state="readonly", justify="right", width=18)
        self.sales_type_filter.pack(side=RIGHT, padx=6)
        rtl_label(filters, "فیلتر محصول").pack(side=RIGHT)
        self.sales_product_filter = ttk.Combobox(filters, state="readonly", justify="right", width=18)
        self.sales_product_filter.pack(side=RIGHT, padx=6)
        self.sales_type_filter.bind("<<ComboboxSelected>>", lambda _event: self.refresh_sales())
        self.sales_product_filter.bind("<<ComboboxSelected>>", lambda _event: self.refresh_sales())
        self.sales_period_notebook = ttk.Notebook(self.sales_tab)
        self.sales_period_notebook.pack(fill=BOTH, expand=True)
        self.sales_trees = {}
        for key, title in [("daily", "روزانه"), ("monthly", "ماهانه"), ("yearly", "سالانه")]:
            frame = tb.Frame(self.sales_period_notebook, padding=8)
            self.sales_period_notebook.add(frame, text=title)
            upper = tb.Frame(frame)
            upper.pack(fill=BOTH, expand=True)
            tree = ttk.Treeview(upper, columns=("period", "type", "product", "count", "revenue"), show="headings")
            for col, text, width in [("period", "بازه", 120), ("type", "نوع معامله", 140), ("product", "محصول", 160), ("count", "تعداد", 90), ("revenue", "درآمد", 140)]:
                tree.heading(col, text=text)
                tree.column(col, width=width, anchor="e")
            tree.pack(side=RIGHT, fill=BOTH, expand=True)
            if not hasattr(self, "sales_charts"):
                self.sales_charts = {}
            chart = LineChart(upper, f"نمودار {title}")
            chart.pack(side=RIGHT, fill=BOTH, expand=True, padx=(10, 0))
            self.sales_trees[key] = tree
            self.sales_charts[key] = chart

    def build_reports(self):
        rtl_label(self.reports_tab, "گزارشات", style="Section.TLabel").pack(fill=X, pady=(0, 10))
        summary_row = tb.Frame(self.reports_tab)
        summary_row.pack(fill=X, pady=(0, 10))
        self.report_commission_total = DashboardCard(summary_row, "جمع کمیسیون اپراتور", "0", "warning")
        self.report_commission_total.pack(side=RIGHT, fill=X, expand=True, padx=6)
        charts = tb.Frame(self.reports_tab)
        charts.pack(fill=BOTH, expand=True)
        self.category_chart = BarChart(charts, "درآمد بر اساس دسته‌بندی")
        self.pipeline_chart = BarChart(charts, "تعداد معامله بر اساس کاریز")
        self.category_chart.pack(side=RIGHT, fill=BOTH, expand=True, padx=6)
        self.pipeline_chart.pack(side=RIGHT, fill=BOTH, expand=True, padx=6)

    def build_notifications(self):
        rtl_label(self.notifications_tab, "اعلان‌ها و یادآورها", style="Section.TLabel").pack(fill=X, pady=(0, 10))
        self.notifications_tree = ttk.Treeview(self.notifications_tab, columns=("title", "customer", "status", "reminder"), show="headings", selectmode="extended")
        for key, title, width in [("title", "معامله", 220), ("customer", "مشتری", 150), ("status", "وضعیت", 130), ("reminder", "زمان یادآور", 180)]:
            self.notifications_tree.heading(key, text=title)
            self.notifications_tree.column(key, width=width, anchor="e")
        self.notifications_tree.pack(fill=BOTH, expand=True)
        self.notifications_tree.bind("<Double-1>", self.open_notification_deal)
        toolbar = tb.Frame(self.notifications_tab)
        toolbar.pack(fill=X, pady=(10, 0))
        tb.Button(toolbar, text="خوانده شد", bootstyle="success-outline", command=self.mark_selected_notifications_read).pack(side=RIGHT, padx=4)
        tb.Button(toolbar, text="همگی خوانده شد", bootstyle="secondary-outline", command=self.mark_all_notifications_read).pack(side=RIGHT, padx=4)
        self.notification_sort = ttk.Combobox(toolbar, state="readonly", justify="right", width=16, values=["جدیدترین", "قدیمی‌ترین", "وضعیت", "عنوان"])
        self.notification_sort.pack(side=RIGHT, padx=6)
        self.notification_sort.set("جدیدترین")
        self.notification_sort.bind("<<ComboboxSelected>>", lambda _event: self.refresh_notifications())

    def build_io(self):
        rtl_label(self.io_tab, "ورودی/خروجی اکسل", style="Section.TLabel").pack(fill=X, pady=(0, 10))
        subtitle = rtl_label(self.io_tab, "برای هر بخش می‌توانید فایل اکسل وارد یا صادر کنید. ستون شناسه لازم نیست و خود برنامه آن را مدیریت می‌کند.", bootstyle="secondary")
        subtitle.pack(fill=X, pady=(0, 12))
        for title in DATASET_CONFIG:
            box = tb.LabelFrame(self.io_tab, text=title)
            box.pack(fill=X, pady=6)
            row = tb.Frame(box, padding=10)
            row.pack(fill=X, pady=6)
            tb.Button(row, text="ایمپورت Xlsx", bootstyle="warning-outline", command=lambda t=title: self.import_dataset(t)).pack(side=RIGHT, padx=4)
            tb.Button(row, text="اکسپورت Xlsx", bootstyle="success-outline", command=lambda t=title: self.export_dataset(t)).pack(side=RIGHT, padx=4)

    def build_guide(self):
        rtl_label(self.guide_tab, "راهنمای جامع نرم‌افزار", style="Section.TLabel").pack(fill=X, pady=(0, 10))

        hero = tb.Frame(self.guide_tab, padding=20, bootstyle="info")
        hero.pack(fill=X, pady=(0, 14))
        rtl_label(hero, "راهنمای شروع سریع و استفاده حرفه‌ای", font=("Tahoma", 14, "bold"), foreground="#0f172a").pack(fill=X, pady=(0, 8))
        rtl_label(
            hero,
            "این بخش برای این طراحی شده که بدون آزمون و خطا، مسیر استفاده از نرم‌افزار را سریع پیدا کنید. همه داده‌ها آفلاین ذخیره می‌شوند و هر کاربر ویندوز می‌تواند پروفایل، مشتریان، معاملات، اعلان‌ها و گزارش‌های مخصوص خودش را داشته باشد.",
            wraplength=1080,
            font=("Tahoma", 10),
            foreground="#0f172a",
        ).pack(fill=X)

        quick_row = tb.Frame(self.guide_tab)
        quick_row.pack(fill=X, pady=(0, 14))
        quick_row.columnconfigure((0, 1, 2), weight=1)
        quick_cards = [
            ("شروع کار", "ساخت پروفایل، ورود امن و آماده‌سازی تنظیمات اولیه"),
            ("مدیریت فروش", "ثبت معامله، پیگیری یادآور، دسته‌بندی و کاریز"),
            ("تحلیل و گزارش", "داشبورد، فروش و درآمد، اعلان‌ها و خروجی اکسل"),
        ]
        for index, (title, text) in enumerate(quick_cards):
            card = tb.Frame(quick_row, padding=14, bootstyle="light")
            card.grid(row=0, column=index, sticky="nsew", padx=6)
            rtl_label(card, title, font=("Tahoma", 11, "bold")).pack(fill=X, pady=(0, 6))
            rtl_label(card, text, wraplength=300, font=("Tahoma", 9)).pack(fill=X)

        sections = [
            ("شروع کار", "برای استفاده از نرم‌افزار، ابتدا یک پروفایل بسازید و با نام کاربری و رمز عبور وارد شوید. بعد از ورود، بهتر است از بخش تنظیمات، دسته‌بندی‌ها، کاریزها، نوع معامله و محصولات را متناسب با روش فروش خودتان تعریف کنید تا ثبت اطلاعات سریع‌تر و دقیق‌تر انجام شود."),
            ("مدیریت مشتریان", "در بخش مشتریان می‌توانید ثبت، ویرایش، حذف و جست‌وجو را بر اساس نام یا شماره موبایل انجام دهید. اگر شماره موبایل تکراری وارد شود، برنامه همان لحظه هشدار می‌دهد و ثبت انجام نمی‌شود. با باز کردن هر مشتری، همه معاملات مرتبط با او در همان پنجره نمایش داده می‌شود."),
            ("مدیریت معاملات", "برای هر معامله می‌توانید عنوان، مشتری، نوع معامله، محصول، دسته‌بندی، کاریز، وضعیت، قیمت فروش، یادداشت و زمان یادآور ثبت کنید. دوبارکلیک روی ردیف معامله، فرم همان معامله را باز می‌کند و از داخل فرم هم می‌توانید سریع به پیام‌رسان‌های مختلف برای ارتباط با مشتری بروید."),
            ("داشبورد", "داشبورد برای مشاهده سریع وضعیت فعلی فروش طراحی شده است. در این بخش خلاصه تعداد مشتریان، تعداد معاملات، موارد در دست بررسی و درآمد ماهیانه دیده می‌شود. همچنین مثل بخش معاملات، جست‌وجو، فیلتر، سورت، چیدمان ستون‌ها و باز کردن معامله از روی لیست هم در دسترس است."),
            ("اعلان‌ها و یادآورها", "اعلان معامله فقط وقتی در لیست اعلان‌ها ثبت می‌شود که زمان یادآور واقعاً فرا رسیده باشد. برای تولد مشتری هم در همان روز اعلان ساخته می‌شود. اعلان‌های بازنشده از اعلان‌های خوانده‌شده جدا هستند و می‌توانید موردی یا گروهی آن‌ها را به وضعیت خوانده‌شده ببرید."),
            ("فروش، درآمد و گزارشات", "در بخش فروش و درآمد، تحلیل روزانه، ماهانه و سالانه با جدول و نمودار نمایش داده می‌شود. گزارشات حرفه‌ای مثل پرفروش‌ترین روز، پرفروش‌ترین ماه، نرخ موفقیت و ترکیب وضعیت معاملات هم برای تحلیل عملکرد تیم فروش در اختیار شما قرار می‌گیرد."),
            ("ورودی و خروجی", "برای مشتریان، معاملات، کاریزها، دسته‌بندی‌ها، نوع معامله و محصولات می‌توانید فایل اکسل با فرمت Xlsx وارد یا صادر کنید. نیازی نیست ستون شناسه در فایل وجود داشته باشد، چون برنامه آن را خودش مدیریت می‌کند. پیشنهاد می‌شود برای پشتیبان‌گیری، به‌صورت دوره‌ای خروجی اکسل تهیه کنید."),
        ]

        body = tb.Frame(self.guide_tab)
        body.pack(fill=BOTH, expand=True)
        body.columnconfigure((0, 1), weight=1)
        for index, (heading, paragraph) in enumerate(sections):
            section_box = tb.Frame(body, padding=16, bootstyle="light")
            section_box.grid(row=index // 2, column=index % 2, sticky="nsew", padx=6, pady=6)
            header = tb.Frame(section_box, bootstyle="secondary")
            header.pack(fill=X, pady=(0, 10))
            rtl_label(header, heading, font=("Tahoma", 11, "bold"), foreground="#111111").pack(fill=X, padx=10, pady=8)
            rtl_label(section_box, paragraph, wraplength=500, font=("Tahoma", 10)).pack(fill=X)

    def build_about(self):
        rtl_label(self.about_tab, "درباره ما", style="Section.TLabel").pack(fill=X, pady=(0, 10))

        intro = tb.Frame(self.about_tab, padding=20, bootstyle="primary")
        intro.pack(fill=X, pady=(0, 14))
        rtl_label(intro, "POWEREN", font=("Tahoma", 15, "bold"), foreground="#ffffff").pack(fill=X, pady=(0, 6))
        rtl_label(
            intro,
            f"این نرم‌افزار به‌صورت آفلاین برای مدیریت حرفه‌ای مشتریان و معاملات ساخته شده و نسخه فعلی آن {APP_VERSION} است و به‌صورت رایگان در اختیار کاربران قرار گرفته است.",
            wraplength=1080,
            font=("Tahoma", 10),
            foreground="#ffffff",
        ).pack(fill=X)

        content = tb.Frame(self.about_tab)
        content.pack(fill=BOTH, expand=True)
        content.columnconfigure((0, 1), weight=1)

        creator = tb.Frame(content, padding=16, bootstyle="light")
        creator.grid(row=0, column=1, sticky="nsew", padx=6, pady=6)
        rtl_label(creator, "مشخصات سازنده", font=("Tahoma", 11, "bold")).pack(fill=X, pady=(0, 8))
        rtl_label(creator, "نام برند: پاوران").pack(fill=X, pady=2)
        rtl_label(creator, "نام لاتین برند: POWEREN").pack(fill=X, pady=2)
        rtl_label(creator, "ایمیل سازنده: siahtirim@gmail.com").pack(fill=X, pady=2)
        rtl_label(creator, "وب‌سایت رسمی: www.poweren.ir").pack(fill=X, pady=2)

        links = tb.Frame(content, padding=16, bootstyle="light")
        links.grid(row=0, column=0, sticky="nsew", padx=6, pady=6)
        rtl_label(links, "لینک‌های ارتباطی و دریافت نسخه جدید", font=("Tahoma", 11, "bold")).pack(fill=X, pady=(0, 8))
        rtl_label(links, "برای دریافت نسخه‌های جدید، مشاهده تغییرات و ارتباط با سازنده از لینک‌های زیر استفاده کنید.").pack(fill=X, pady=(0, 10))

        rtl_label(links, "وب‌سایت", font=("Tahoma", 10, "bold")).pack(fill=X)
        site = tb.Label(links, text="https://www.poweren.ir", foreground="#0d6efd", cursor="hand2", anchor="e", justify="right")
        site.pack(fill=X, pady=(2, 10))
        site.bind("<Button-1>", lambda _event: webbrowser.open("https://www.poweren.ir"))

        rtl_label(links, "گیت‌هاب - دریافت آپدیت جدید", font=("Tahoma", 10, "bold")).pack(fill=X)
        github = tb.Label(links, text="https://github.com/siahtirilab/Offline-CRM", foreground="#0d6efd", cursor="hand2", anchor="e", justify="right")
        github.pack(fill=X, pady=(2, 10))
        github.bind("<Button-1>", lambda _event: webbrowser.open("https://github.com/siahtirilab/Offline-CRM"))

        rtl_label(links, "ایمیل", font=("Tahoma", 10, "bold")).pack(fill=X)
        email = tb.Label(links, text="mailto:siahtirim@gmail.com", foreground="#0d6efd", cursor="hand2", anchor="e", justify="right")
        email.pack(fill=X)
        email.bind("<Button-1>", lambda _event: webbrowser.open("mailto:siahtirim@gmail.com"))

    def build_settings(self):
        options_box = tb.LabelFrame(self.settings_tab, text="تنظیمات عمومی")
        options_box.pack(fill=X, pady=(0, 12))
        options_row = tb.Frame(options_box, padding=12)
        options_row.pack(fill=X)
        tb.Checkbutton(
            options_row,
            text="ثبت برنامه در استارت‌آپ ویندوز",
            variable=self.startup_enabled_var,
            bootstyle="success-round-toggle",
            command=self.toggle_startup_setting,
        ).pack(side=RIGHT)
        wrapper = tb.Frame(self.settings_tab)
        wrapper.pack(fill=BOTH, expand=True)
        self.category_manager = CategoryManager(wrapper, self.store.categories(), self.save_categories)
        self.category_manager.pack(side=RIGHT, fill=BOTH, expand=True, padx=6)
        self.pipeline_manager = TextListManager(wrapper, "کاریزها", self.store.pipelines(), self.save_pipelines)
        self.pipeline_manager.pack(side=RIGHT, fill=BOTH, expand=True, padx=6)
        bottom = tb.Frame(self.settings_tab)
        bottom.pack(fill=BOTH, expand=True, pady=(12, 0))
        self.deal_type_manager = TextListManager(bottom, "نوع معامله", self.store.deal_types(), self.save_deal_types)
        self.deal_type_manager.pack(side=RIGHT, fill=BOTH, expand=True, padx=6)
        self.product_manager = ProductManager(bottom, self.store.deal_types(), self.store.products(), self.save_products)
        self.product_manager.pack(side=RIGHT, fill=BOTH, expand=True, padx=6)

    def refresh_everything(self):
        self._customers = self.store.customers()
        self._deals = self.store.deals()
        self._categories = self.store.categories()
        self._pipelines = self.store.pipelines()
        self._deal_types = self.store.deal_types()
        self._products = self.store.products()
        deal_status_options = ["همه وضعیت‌ها"] + DEAL_STATUSES
        for combo in [self.deal_status_filter, self.dashboard_status_filter]:
            current = combo.get()
            combo.configure(values=deal_status_options)
            combo.set(current if current in deal_status_options else "همه وضعیت‌ها")
        deal_category_options = ["همه دسته‌بندی‌ها"] + [item["title"] for item in self._categories]
        current_category = self.deal_category_filter.get()
        self.deal_category_filter.configure(values=deal_category_options)
        self.deal_category_filter.set(current_category if current_category in deal_category_options else "همه دسته‌بندی‌ها")
        sales_type_options = ["همه"] + [item["title"] for item in self._deal_types]
        current_type = self.sales_type_filter.get()
        self.sales_type_filter.configure(values=sales_type_options)
        self.sales_type_filter.set(current_type if current_type in sales_type_options else "همه")
        current_dashboard_type = self.dashboard_type_filter.get()
        self.dashboard_type_filter.configure(values=sales_type_options)
        self.dashboard_type_filter.set(current_dashboard_type if current_dashboard_type in sales_type_options else "همه")
        current_deal_type = self.deal_type_filter.get()
        self.deal_type_filter.configure(values=sales_type_options)
        self.deal_type_filter.set(current_deal_type if current_deal_type in sales_type_options else "همه")
        sales_product_options = ["همه"] + [item["title"] for item in self._products]
        current_product = self.sales_product_filter.get()
        self.sales_product_filter.configure(values=sales_product_options)
        self.sales_product_filter.set(current_product if current_product in sales_product_options else "همه")
        current_dashboard_product = self.dashboard_product_filter.get()
        self.dashboard_product_filter.configure(values=sales_product_options)
        self.dashboard_product_filter.set(current_dashboard_product if current_dashboard_product in sales_product_options else "همه")
        current_deal_product = self.deal_product_filter.get()
        self.deal_product_filter.configure(values=sales_product_options)
        self.deal_product_filter.set(current_deal_product if current_deal_product in sales_product_options else "همه")
        self.refresh_customer_tree()
        self.refresh_deal_tree()
        self.refresh_dashboard_filter()
        self.refresh_dashboard()
        self.refresh_sales()
        self.refresh_reports()
        self.refresh_notifications()
        self.category_manager.rows = [dict(item) for item in self._categories]
        self.category_manager.refresh()
        self.pipeline_manager.rows = [dict(item) for item in self._pipelines]
        self.pipeline_manager.refresh()
        self.deal_type_manager.rows = [dict(item) for item in self._deal_types]
        self.deal_type_manager.refresh()
        self.product_manager.rows = [dict(item) for item in self._products]
        self.product_manager.set_deal_types(self._deal_types)
        self.product_manager.refresh()

    def refresh_customer_tree(self):
        for item in self.customer_tree.get_children():
            self.customer_tree.delete(item)
        query = self.customer_search.get().strip() if hasattr(self, "customer_search") else ""
        rows = []
        for customer in self._customers:
            if not matches_search(query, customer["name"], customer["mobile"]):
                continue
            rows.append(customer)
        sort_key = self.customer_sort.get() if hasattr(self, "customer_sort") else "بدون سورت"
        if sort_key == "نام":
            rows = sorted(rows, key=lambda item: item.get("name", ""))
        elif sort_key == "موبایل":
            rows = sorted(rows, key=lambda item: item.get("mobile", ""))
        elif sort_key == "تاریخ تولد":
            rows = sorted(rows, key=lambda item: normalize_jalali_date(item.get("birthdate", "")) or "9999/99/99")
        elif sort_key == "سطح درآمد":
            order = {level: index for index, level in enumerate(INCOME_LEVELS)}
            rows = sorted(rows, key=lambda item: order.get(item.get("income_level", ""), 999))
        page_rows, total_pages = self.paginate_rows(rows, self.customer_page, self.current_page_size(self.customer_tree))
        self.customer_page = min(max(1, self.customer_page), total_pages)
        if hasattr(self, "customer_page_label"):
            self.customer_page_label.configure(text=f"صفحه {self.customer_page} از {total_pages}")
        if hasattr(self, "customer_total_label"):
            self.customer_total_label.configure(text=f"تعداد کل: {len(rows)}")
        if hasattr(self, "customer_page_entry"):
            self.customer_page_entry.delete(0, END)
            self.customer_page_entry.insert(0, str(self.customer_page))
        for customer in page_rows:
            self.customer_tree.insert("", END, iid=customer["id"], values=(customer["name"], customer["mobile"], customer["birthdate"], customer["income_level"]))

    def refresh_deal_tree(self):
        for item in self.deal_tree.get_children():
            self.deal_tree.delete(item)
        category_color_map = {row["title"]: row.get("color", "") for row in self._categories}
        configured_tags = set()
        search = self.deal_search.get().strip() if hasattr(self, "deal_search") else ""
        rows = []
        for deal in self.filtered_deals_for_page():
            if not matches_search(search, deal["title"], deal["customer_name"], deal.get("product", ""), deal.get("notes", "")):
                continue
            rows.append(deal)
        page_rows, total_pages = self.paginate_rows(rows, self.deal_page, self.current_page_size(self.deal_tree))
        self.deal_page = min(max(1, self.deal_page), total_pages)
        if hasattr(self, "deal_page_label"):
            self.deal_page_label.configure(text=f"صفحه {self.deal_page} از {total_pages}")
        if hasattr(self, "deal_total_label"):
            self.deal_total_label.configure(text=f"تعداد کل: {len(rows)}")
        if hasattr(self, "deal_page_entry"):
            self.deal_page_entry.delete(0, END)
            self.deal_page_entry.insert(0, str(self.deal_page))
        for deal in page_rows:
            status_label = deal["status"]
            category_label = category_display(deal["category"])
            category_tag = f"deal_category_{deal.get('category', '').strip() or 'default'}"
            if category_tag not in configured_tags:
                self.deal_tree.tag_configure(category_tag, background=soften_hex_color(category_color_map.get(deal["category"], "")), foreground="#111111")
                configured_tags.add(category_tag)
            self.deal_tree.insert(
                "",
                END,
                iid=deal["id"],
                values=(deal["title"], deal["customer_name"], deal["deal_type"], deal.get("product", ""), category_label, deal["pipeline"], status_label, deal["sale_price"], (deal.get("created_at", "").split(" ")[0]), deal["reminder_at"], note_preview_text(deal.get("notes", ""))),
                tags=(category_tag,),
            )
        for item in self.dashboard_deal_tree.get_children():
            self.dashboard_deal_tree.delete(item)
        for deal in self.filtered_deals():
            status_label = deal["status"]
            category_label = category_display(deal["category"])
            category_tag = f"dashboard_category_{deal.get('category', '').strip() or 'default'}"
            if category_tag not in configured_tags:
                self.dashboard_deal_tree.tag_configure(category_tag, background=soften_hex_color(category_color_map.get(deal["category"], "")), foreground="#111111")
                configured_tags.add(category_tag)
            self.dashboard_deal_tree.insert(
                "",
                END,
                iid=deal["id"],
                values=(deal["title"], deal["customer_name"], status_label, category_label, (deal.get("created_at", "").split(" ")[0]), deal["reminder_at"]),
                tags=(category_tag,),
            )

    def deal_tree_column_name(self, column_id: str) -> str:
        if not column_id.startswith("#"):
            return ""
        try:
            index = int(column_id[1:]) - 1
        except ValueError:
            return ""
        display_columns = list(self.deal_tree["displaycolumns"])
        if 0 <= index < len(display_columns):
            return str(display_columns[index])
        return ""

    def on_deal_tree_motion(self, event):
        row_id = self.deal_tree.identify_row(event.y)
        column_id = self.deal_tree.identify_column(event.x)
        if not row_id or self.deal_tree_column_name(column_id) != "notes":
            self.deal_notes_tooltip.hide()
            return
        deal = next((item for item in self._deals if item["id"] == row_id), None)
        note_text = (deal or {}).get("notes", "").strip()
        if not note_text:
            self.deal_notes_tooltip.hide()
            return
        tooltip_text = "\n".join(line for line in note_text.splitlines() if line.strip())
        self.deal_notes_tooltip.show(tooltip_text, event.x_root + 18, event.y_root + 18)

    def refresh_dashboard_filter(self):
        options = ["همه دسته‌بندی‌ها"] + [item["title"] for item in self._categories]
        current = self.dashboard_category_filter.get() if hasattr(self, "dashboard_category_filter") else ""
        self.dashboard_category_filter.configure(values=options)
        self.dashboard_category_filter.set(current if current in options else options[0])

    def filtered_deals(self) -> list[dict]:
        deals = list(self._deals)
        search = self.dashboard_search.get().strip() if hasattr(self, "dashboard_search") else ""
        if search:
            deals = [
                deal
                for deal in deals
                if matches_search(search, deal.get("title", ""), deal.get("customer_name", ""), deal.get("product", ""), deal.get("notes", ""))
            ]
        selected = self.dashboard_category_filter.get()
        if selected and selected != "همه دسته‌بندی‌ها":
            deals = [deal for deal in deals if deal["category"] == selected]
        status_filter = self.dashboard_status_filter.get()
        if status_filter and status_filter != "همه وضعیت‌ها":
            deals = [deal for deal in deals if deal["status"] == status_filter]
        product_filter = self.dashboard_product_filter.get()
        if product_filter and product_filter != "همه":
            deals = [deal for deal in deals if deal.get("product", "") == product_filter]
        type_filter = self.dashboard_type_filter.get()
        if type_filter and type_filter != "همه":
            deals = [deal for deal in deals if deal["deal_type"] == type_filter]
        return self.sort_deals(deals, self.dashboard_sort.get())

    def filtered_deals_for_page(self) -> list[dict]:
        deals = list(self._deals)
        status_filter = self.deal_status_filter.get()
        if status_filter and status_filter != "همه وضعیت‌ها":
            deals = [deal for deal in deals if deal["status"] == status_filter]
        category_filter = self.deal_category_filter.get()
        if category_filter and category_filter != "همه دسته‌بندی‌ها":
            deals = [deal for deal in deals if deal["category"] == category_filter]
        product_filter = self.deal_product_filter.get()
        if product_filter and product_filter != "همه":
            deals = [deal for deal in deals if deal.get("product", "") == product_filter]
        type_filter = self.deal_type_filter.get()
        if type_filter and type_filter != "همه":
            deals = [deal for deal in deals if deal["deal_type"] == type_filter]
        reminder_day = self.deal_reminder_filter.get().strip() if hasattr(self, "deal_reminder_filter") else ""
        if reminder_day:
            deals = [deal for deal in deals if normalize_jalali_date((deal.get("reminder_at", "").split(" ") or [""])[0]) == reminder_day]
        return self.sort_deals(deals, self.deal_sort.get())

    def sort_deals(self, deals: list[dict], sort_key: str) -> list[dict]:
        key_map = {
            "یادآور": lambda item: parse_jalali_datetime(item.get("reminder_at", "")) or datetime.max,
            "وضعیت": lambda item: item.get("status", ""),
            "دسته‌بندی": lambda item: item.get("category", ""),
            "محصول": lambda item: item.get("product", ""),
            "نوع معامله": lambda item: item.get("deal_type", ""),
        }
        if sort_key == "یادآور":
            return sorted(deals, key=key_map[sort_key], reverse=True)
        if sort_key in key_map:
            return sorted(deals, key=key_map[sort_key])
        return deals

    def paginate_rows(self, rows: list[dict], current_page: int, page_size: int | None = None) -> tuple[list[dict], int]:
        page_size = max(1, page_size or self.list_page_size)
        total_pages = max(1, (len(rows) + page_size - 1) // page_size)
        current_page = min(max(1, current_page), total_pages)
        start = (current_page - 1) * page_size
        end = start + page_size
        return rows[start:end], total_pages

    def change_customer_page(self, step: int):
        self.customer_page = max(1, self.customer_page + step)
        self.refresh_customer_tree()

    def change_deal_page(self, step: int):
        self.deal_page = max(1, self.deal_page + step)
        self.refresh_deal_tree()

    def go_customer_page(self):
        try:
            self.customer_page = max(1, int(self.customer_page_entry.get().strip() or "1"))
        except ValueError:
            self.customer_page = 1
        self.refresh_customer_tree()

    def go_deal_page(self):
        try:
            self.deal_page = max(1, int(self.deal_page_entry.get().strip() or "1"))
        except ValueError:
            self.deal_page = 1
        self.refresh_deal_tree()

    def reset_customer_page(self):
        self.customer_page = 1
        self.refresh_customer_tree()

    def reset_deal_page(self):
        self.deal_page = 1
        self.refresh_deal_tree()

    def mark_selected_notifications_read(self):
        selected = list(self.notifications_tree.selection())
        if not selected:
            Messagebox.show_warning("ابتدا یک یا چند اعلان را انتخاب کنید.", "اعلان‌ها", parent=self)
            return
        self._mark_notification_ids_as_read(selected)

    def mark_all_notifications_read(self):
        unread = []
        for item_id in self.notifications_tree.get_children():
            tags = self.notifications_tree.item(item_id, "tags")
            if "unread" in tags:
                unread.append(item_id)
        if not unread:
            Messagebox.show_info("اعلان خوانده‌نشده‌ای وجود ندارد.", "اعلان‌ها", parent=self)
            return
        self._mark_notification_ids_as_read(unread)

    def _mark_notification_ids_as_read(self, item_ids: list[str]):
        deal_ids = []
        birthday_ids = []
        for item_id in item_ids:
            kind, record_id = item_id.split(":", 1)
            if kind == "deal":
                deal_ids.append(record_id)
            elif kind == "birthday":
                birthday_ids.append(record_id)
        if deal_ids:
            rows = self.store.deals()
            changed = False
            for row in rows:
                if row["id"] in deal_ids:
                    row["notification_seen_at"] = row.get("reminder_at", "")
                    changed = True
            if changed:
                self.store.save_deals(rows)
                self._deals = rows
        if birthday_ids:
            now = jdatetime.datetime.fromgregorian(datetime=datetime.now())
            today_key = f"{now.year:04d}/{now.month:02d}/{now.day:02d}"
            rows = self.store.customers()
            changed = False
            for row in rows:
                if row["id"] in birthday_ids:
                    row["birthday_seen_for"] = today_key
                    changed = True
            if changed:
                self.store.save_customers(rows)
                self._customers = rows
        self.refresh_notifications()

    def delete_dashboard_deal(self):
        deal = self.selected_deal_from_tree(self.dashboard_deal_tree)
        if not deal:
            Messagebox.show_warning("ابتدا یک معامله را در داشبورد انتخاب کنید.", "داشبورد", parent=self)
            return
        if Messagebox.yesno("معامله انتخاب شده حذف شود؟", "حذف معامله", parent=self) != "Yes":
            return
        self.store.save_deals([row for row in self.store.deals() if row["id"] != deal["id"]])
        self.refresh_everything()

    def refresh_dashboard(self):
        deals = self.filtered_deals()
        success_deals = [deal for deal in deals if deal["status"] == "موفق"]
        in_review_deals = [deal for deal in deals if deal["status"] == "در دست بررسی"]
        current_month = jdatetime.datetime.now().strftime("%Y/%m")
        monthly_commission = 0.0
        for deal in success_deals:
            created = deal.get("created_at", "")
            if created.startswith(current_month):
                monthly_commission += to_float(deal.get("operator_commission", ""))
        self.card_customers.set_value(str(len(self._customers)))
        self.card_deals.set_value(str(len(deals)))
        self.card_success.set_value(str(len(in_review_deals)))
        self.card_revenue.set_value(f"{monthly_commission:,.0f}")

        for item in self.dashboard_deal_tree.get_children():
            self.dashboard_deal_tree.delete(item)
        category_color_map = {row["title"]: row.get("color", "") for row in self._categories}
        configured_tags = set()
        for deal in deals:
            category_tag = f"dashboard_category_{deal.get('category', '').strip() or 'default'}"
            if category_tag not in configured_tags:
                self.dashboard_deal_tree.tag_configure(category_tag, background=soften_hex_color(category_color_map.get(deal["category"], "")), foreground="#111111")
                configured_tags.add(category_tag)
            self.dashboard_deal_tree.insert(
                "",
                END,
                iid=deal["id"],
                values=(deal["title"], deal["customer_name"], deal["status"], category_display(deal["category"]), (deal.get("created_at", "").split(" ")[0]), deal["reminder_at"]),
                tags=(category_tag,),
            )

    def refresh_reports(self):
        success_deals = [deal for deal in self._deals if deal["status"] == "موفق"]
        revenue_by_category = defaultdict(float)
        count_by_pipeline = defaultdict(float)
        total_commission = 0.0
        for deal in success_deals:
            revenue_by_category[deal["category"] or "بدون دسته"] += to_float(deal["sale_price"])
            total_commission += to_float(deal.get("operator_commission", ""))
        for deal in self._deals:
            count_by_pipeline[deal["pipeline"] or "بدون کاریز"] += 1
        self.report_commission_total.set_value(f"{total_commission:,.0f}")
        self.category_chart.render(sorted(revenue_by_category.items(), key=lambda item: item[1], reverse=True)[:7])
        self.pipeline_chart.render(sorted(count_by_pipeline.items(), key=lambda item: item[1], reverse=True)[:7])

    def refresh_sales(self):
        all_deals = list(self._deals)
        success_deals = [deal for deal in all_deals if deal["status"] == "موفق"]
        selected_type = self.sales_type_filter.get()
        selected_product = self.sales_product_filter.get()
        if selected_type and selected_type != "همه":
            all_deals = [deal for deal in all_deals if deal["deal_type"] == selected_type]
        if selected_type and selected_type != "همه":
            success_deals = [deal for deal in success_deals if deal["deal_type"] == selected_type]
        if selected_product and selected_product != "همه":
            all_deals = [deal for deal in all_deals if deal.get("product", "") == selected_product]
        if selected_product and selected_product != "همه":
            success_deals = [deal for deal in success_deals if deal.get("product", "") == selected_product]
        total_revenue = sum(to_float(deal["sale_price"]) for deal in success_deals)
        self.sales_summary.configure(text=f"جمع فروش موفق: {total_revenue:,.0f}")
        daily_revenue = defaultdict(float)
        monthly_revenue = defaultdict(float)
        status_count = defaultdict(int)
        for deal in all_deals:
            dt = parse_jalali_datetime(deal.get("created_at", ""))
            jalali_dt = jdatetime.datetime.fromgregorian(datetime=dt) if dt else jdatetime.datetime.now()
            day_key = f"{jalali_dt.year:04d}/{jalali_dt.month:02d}/{jalali_dt.day:02d}"
            month_key = f"{jalali_dt.year:04d}/{jalali_dt.month:02d}"
            if deal["status"] == "موفق":
                daily_revenue[day_key] += to_float(deal["sale_price"])
                monthly_revenue[month_key] += to_float(deal["sale_price"])
            status_count[deal["status"]] += 1
        best_day = max(daily_revenue.items(), key=lambda item: item[1], default=("-", 0))
        best_month = max(monthly_revenue.items(), key=lambda item: item[1], default=("-", 0))
        success_rate = (len(success_deals) / len(all_deals) * 100) if all_deals else 0
        self.sales_best_day.set_value(f"{best_day[0]} | {best_day[1]:,.0f}" if best_day[0] != "-" else "-")
        self.sales_best_month.set_value(f"{best_month[0]} | {best_month[1]:,.0f}" if best_month[0] != "-" else "-")
        self.sales_success_rate.set_value(f"{success_rate:.1f}%")
        self.sales_status_mix.set_value(f"م:{status_count['موفق']} ن:{status_count['ناموفق']} ب:{status_count['در دست بررسی']}")
        aggregations = {"daily": defaultdict(lambda: {"count": 0, "revenue": 0.0}), "monthly": defaultdict(lambda: {"count": 0, "revenue": 0.0}), "yearly": defaultdict(lambda: {"count": 0, "revenue": 0.0})}
        for deal in success_deals:
            dt = parse_jalali_datetime(deal.get("updated_at", "") or deal.get("created_at", ""))
            jalali_dt = jdatetime.datetime.fromgregorian(datetime=dt) if dt else jdatetime.datetime.now()
            keys = {
                "daily": f"{jalali_dt.year:04d}/{jalali_dt.month:02d}/{jalali_dt.day:02d}",
                "monthly": f"{jalali_dt.year:04d}/{jalali_dt.month:02d}",
                "yearly": f"{jalali_dt.year:04d}",
            }
            for period_key, label in keys.items():
                bucket = aggregations[period_key][(label, deal["deal_type"] or "نامشخص", deal.get("product", "") or "بدون محصول")]
                bucket["count"] += 1
                bucket["revenue"] += to_float(deal["sale_price"])
        for period_key, tree in self.sales_trees.items():
            for item in tree.get_children():
                tree.delete(item)
            rows = sorted(aggregations[period_key].items(), key=lambda item: item[1]["revenue"], reverse=True)
            for (period, deal_type, product), metrics in rows:
                tree.insert("", END, values=(period, deal_type, product, int(metrics["count"]), f"{metrics['revenue']:,.0f}"))
            self.sales_charts[period_key].render([(f"{period}\n{product}", metrics["revenue"]) for (period, _deal_type, product), metrics in rows[:7]])

    def refresh_notifications(self):
        for item in self.notifications_tree.get_children():
            self.notifications_tree.delete(item)
        now = datetime.now()
        due_count = 0
        rows = []
        for deal in sorted(self._deals, key=lambda item: item.get("reminder_at", "")):
            due = parse_jalali_datetime(deal.get("reminder_at", ""))
            if due is None:
                continue
            if deal.get("last_notified_at", "") != deal.get("reminder_at", ""):
                continue
            if due <= now:
                due_count += 1
            rows.append({
                "iid": f"deal:{deal['id']}",
                "title": deal["title"],
                "customer": deal["customer_name"],
                "status": deal["status"],
                "reminder": deal["reminder_at"],
                "read": deal.get("notification_seen_at", "") == deal.get("reminder_at", ""),
            })
        jalali_today = jdatetime.datetime.fromgregorian(datetime=now)
        today_key = f"{jalali_today.year:04d}/{jalali_today.month:02d}/{jalali_today.day:02d}"
        for customer in self._customers:
            birthdate = normalize_jalali_date(customer.get("birthdate", ""))
            if not birthdate:
                continue
            try:
                birth = jdatetime.datetime.strptime(birthdate, "%Y/%m/%d")
            except ValueError:
                continue
            if birth.month == jalali_today.month and birth.day == jalali_today.day and customer.get("birthday_notified_for", "") == today_key:
                rows.append({
                    "iid": f"birthday:{customer['id']}",
                    "title": "تولد مشتری",
                    "customer": customer["name"],
                    "status": "تولد",
                    "reminder": f"{today_key} 10:00",
                    "read": customer.get("birthday_seen_for", "") == today_key,
                })
        sort_key = self.notification_sort.get() if hasattr(self, "notification_sort") else "جدیدترین"
        key_map = {
            "یادآور": lambda item: parse_jalali_datetime(item["reminder"]) or datetime.max,
            "وضعیت": lambda item: item["status"],
            "عنوان": lambda item: item["title"],
        }
        if sort_key == "جدیدترین":
            rows = sorted(rows, key=lambda item: parse_jalali_datetime(item["reminder"]) or datetime.min, reverse=True)
        elif sort_key == "قدیمی‌ترین":
            rows = sorted(rows, key=lambda item: parse_jalali_datetime(item["reminder"]) or datetime.max)
        elif sort_key in key_map:
            rows = sorted(rows, key=key_map[sort_key])
        unread_count = 0
        for row in rows:
            if not row["read"]:
                unread_count += 1
            tag = "read" if row["read"] else "unread"
            self.notifications_tree.tag_configure("read", background="#eef2f3")
            self.notifications_tree.tag_configure("unread", background="#fff4d6")
            self.notifications_tree.insert("", END, iid=row["iid"], values=(row["title"], row["customer"], row["status"], row["reminder"]), tags=(tag,))
        self.notification_badge.configure(text=rtl_text(f"اعلان‌های نخوانده: {unread_count}"))

    def save_categories(self, rows: list[dict]):
        self.store.save_categories(rows)
        self.refresh_everything()

    def save_pipelines(self, rows: list[dict]):
        self.store.save_pipelines(rows)
        self.refresh_everything()

    def save_deal_types(self, rows: list[dict]):
        self.store.save_deal_types(rows)
        valid_types = {row["title"] for row in rows}
        filtered_products = [item for item in self.store.products() if item["deal_type"] in valid_types]
        self.store.save_products(filtered_products)
        self.refresh_everything()

    def save_products(self, rows: list[dict]):
        self.store.save_products(rows)
        self.refresh_everything()

    def export_dataset(self, title: str):
        filename, fields = DATASET_CONFIG[title]
        export_fields = [field for field in fields if field != "id"]
        path = filedialog.asksaveasfilename(parent=self, defaultextension=".xlsx", filetypes=[("Excel", "*.xlsx")], initialfile=filename.replace(".csv", ".xlsx"))
        if not path:
            return
        if title == "مشتریان":
            rows = self.store.customers()
        elif title == "معاملات":
            rows = self.store.deals()
        elif title == "کاریزها":
            rows = self.store.pipelines()
        elif title == "دسته‌بندی‌های معامله":
            rows = self.store.categories()
        elif title == "نوع معامله":
            rows = self.store.deal_types()
        else:
            rows = self.store.products()
        wb = Workbook()
        ws = wb.active
        ws.append(export_fields)
        for row in rows:
            ws.append([row.get(field, "") for field in export_fields])
        wb.save(path)
        Messagebox.show_info("فایل اکسل با موفقیت ذخیره شد.", "اکسپورت", parent=self)

    def import_dataset(self, title: str):
        filename, fields = DATASET_CONFIG[title]
        path = filedialog.askopenfilename(parent=self, filetypes=[("Excel", "*.xlsx")])
        if not path:
            return
        wb = load_workbook(path)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return
        header = [str(cell or "") for cell in rows[0]]
        mapped_rows = []
        for values in rows[1:]:
            row = {"id": uuid.uuid4().hex}
            for field in fields:
                if field == "id":
                    continue
                idx = header.index(field) if field in header else -1
                row[field] = str(values[idx]) if idx >= 0 and idx < len(values) and values[idx] is not None else ""
            mapped_rows.append(row)
        if title == "مشتریان":
            self.store.save_customers(mapped_rows)
        elif title == "معاملات":
            self.store.save_deals(mapped_rows)
        elif title == "کاریزها":
            self.store.save_pipelines(mapped_rows)
        elif title == "دسته‌بندی‌های معامله":
            self.store.save_categories(mapped_rows)
        elif title == "نوع معامله":
            self.store.save_deal_types(mapped_rows)
        else:
            self.store.save_products(mapped_rows)
        self.refresh_everything()
        Messagebox.show_info("ورود اطلاعات از اکسل انجام شد.", "ایمپورت", parent=self)

    def selected_customer(self) -> dict | None:
        selected = self.customer_tree.selection()
        return next((item for item in self._customers if selected and item["id"] == selected[0]), None)

    def open_customer_by_id(self, customer_id: str):
        customer = next((item for item in self._customers if item["id"] == customer_id), None)
        if not customer:
            return
        self.notebook.select(self.customers_tab)
        self.customer_tree.selection_set(customer_id)
        self.customer_tree.focus(customer_id)
        self.customer_tree.see(customer_id)
        related_deals = [deal for deal in self._deals if deal.get("customer_id") == customer_id]
        CustomerDialog(self, customer, related_deals, self.open_deal_by_id, [row["mobile"] for row in self.store.customers()])

    def selected_deal(self) -> dict | None:
        selected = self.deal_tree.selection()
        return next((item for item in self._deals if selected and item["id"] == selected[0]), None)

    def selected_deal_from_tree(self, tree) -> dict | None:
        selected = tree.selection()
        return next((item for item in self._deals if selected and item["id"] == selected[0]), None)

    def add_customer(self):
        dialog = CustomerDialog(self, existing_mobiles=[row["mobile"] for row in self.store.customers()])
        if not dialog.result:
            return
        rows = self.store.customers()
        rows.append(
            {
                "id": uuid.uuid4().hex,
                "name": dialog.result["name"],
                "mobile": dialog.result["mobile"],
                "birthdate": dialog.result["birthdate"],
                "income_level": dialog.result["income_level"],
                "birthday_notified_for": "",
                "birthday_seen_for": "",
                "created_at": now_text(),
                "updated_at": now_text(),
            }
        )
        self.store.save_customers(rows)
        self.refresh_everything()

    def edit_customer(self):
        customer = self.selected_customer()
        if not customer:
            Messagebox.show_warning("ابتدا یک مشتری را انتخاب کنید.", "مشتری", parent=self)
            return
        related_deals = [deal for deal in self._deals if deal.get("customer_id") == customer["id"]]
        dialog = CustomerDialog(self, customer, related_deals, self.open_deal_by_id, [row["mobile"] for row in self.store.customers()])
        if not dialog.result:
            return
        rows = self.store.customers()
        for row in rows:
            if row["id"] == customer["id"]:
                previous_birthdate = row.get("birthdate", "")
                previous_birthday_flag = row.get("birthday_notified_for", "")
                row.update(dialog.result)
                if dialog.result.get("birthdate", "") == previous_birthdate:
                    row["birthday_notified_for"] = previous_birthday_flag
                row["updated_at"] = now_text()
        self.store.save_customers(rows)
        self.refresh_everything()

    def delete_customer(self):
        customer = self.selected_customer()
        if not customer:
            Messagebox.show_warning("ابتدا یک مشتری را انتخاب کنید.", "مشتری", parent=self)
            return
        if Messagebox.yesno("مشتری انتخاب شده حذف شود؟", "حذف مشتری", parent=self) != "Yes":
            return
        self.store.save_customers([row for row in self.store.customers() if row["id"] != customer["id"]])
        self.store.save_deals([row for row in self.store.deals() if row["customer_id"] != customer["id"]])
        self.refresh_everything()

    def add_deal(self):
        if not self._customers:
            Messagebox.show_warning("ابتدا حداقل یک مشتری ثبت کنید.", "معامله", parent=self)
            return
        dialog = DealDialog(self, self._customers, self._categories, self._pipelines, self._deal_types, self._products, existing_deals=self._deals)
        if not dialog.result:
            return
        customer = next((item for item in self._customers if item["name"] == dialog.result["customer_name"]), None)
        reminder_dt = parse_jalali_datetime(dialog.result["reminder_at"])
        rows = self.store.deals()
        rows.append(
            {
                "id": uuid.uuid4().hex,
                "customer_id": customer["id"] if customer else "",
                "customer_name": dialog.result["customer_name"],
                "title": dialog.result["title"],
                "deal_type": dialog.result["deal_type"],
                "product": dialog.result.get("product", ""),
                "category": dialog.result["category"],
                "pipeline": dialog.result["pipeline"],
                "status": dialog.result["status"],
                "sale_price": dialog.result["sale_price"],
                "operator_commission": dialog.result.get("operator_commission", ""),
                "sales_expert": dialog.result["sales_expert"],
                "notes": dialog.result["notes"],
                "reminder_at": dialog.result["reminder_at"],
                "last_notified_at": dialog.result["reminder_at"] if reminder_dt and reminder_dt <= datetime.now() else "",
                "notification_seen_at": "",
                "created_at": now_text(),
                "updated_at": now_text(),
            }
        )
        self.store.save_deals(rows)
        self.refresh_everything()

    def edit_deal(self):
        deal = self.selected_deal()
        self._edit_deal_object(deal)

    def _edit_deal_object(self, deal: dict | None):
        if not deal:
            Messagebox.show_warning("ابتدا یک معامله را انتخاب کنید.", "معامله", parent=self)
            return
        dialog = DealDialog(self, self._customers, self._categories, self._pipelines, self._deal_types, self._products, existing_deals=self._deals, deal=deal)
        if not dialog.result:
            return
        customer = next((item for item in self._customers if item["name"] == dialog.result["customer_name"]), None)
        reminder_dt = parse_jalali_datetime(dialog.result["reminder_at"])
        rows = self.store.deals()
        for row in rows:
            if row["id"] == deal["id"]:
                row.update(dialog.result)
                row["customer_id"] = customer["id"] if customer else ""
                row["last_notified_at"] = dialog.result["reminder_at"] if reminder_dt and reminder_dt <= datetime.now() else ""
                row["updated_at"] = now_text()
        self.store.save_deals(rows)
        self.refresh_everything()

    def edit_selected_deal_from(self, tree):
        self._edit_deal_object(self.selected_deal_from_tree(tree))

    def open_notification_deal(self, event=None):
        item_id = ""
        if event is not None:
            item_id = self.notifications_tree.identify_row(event.y)
            if item_id:
                self.notifications_tree.selection_set(item_id)
        if not item_id:
            selected = self.notifications_tree.selection()
            if not selected:
                return
            item_id = selected[0]
        kind, record_id = item_id.split(":", 1)
        self.open_notification_target(kind, record_id)

    def open_deal_by_id(self, deal_id: str):
        deal = next((item for item in self._deals if item["id"] == deal_id), None)
        if not deal:
            return
        self.notebook.select(self.deals_tab)
        self.deal_tree.selection_set(deal["id"])
        self.deal_tree.focus(deal["id"])
        self.deal_tree.see(deal["id"])
        self._edit_deal_object(deal)

    def mark_deal_notification_seen(self, deal_id: str):
        rows = self.store.deals()
        changed = False
        for row in rows:
            if row["id"] == deal_id:
                row["notification_seen_at"] = row.get("reminder_at", "")
                changed = True
        if changed:
            self.store.save_deals(rows)
            self._deals = rows

    def mark_birthday_seen(self, customer_id: str):
        rows = self.store.customers()
        now = jdatetime.datetime.fromgregorian(datetime=datetime.now())
        today_key = f"{now.year:04d}/{now.month:02d}/{now.day:02d}"
        changed = False
        for row in rows:
            if row["id"] == customer_id:
                row["birthday_seen_for"] = today_key
                changed = True
        if changed:
            self.store.save_customers(rows)
            self._customers = rows

    def delete_deal(self):
        deal = self.selected_deal()
        if not deal:
            Messagebox.show_warning("ابتدا یک معامله را انتخاب کنید.", "معامله", parent=self)
            return
        if Messagebox.yesno("معامله انتخاب شده حذف شود؟", "حذف معامله", parent=self) != "Yes":
            return
        self.store.save_deals([row for row in self.store.deals() if row["id"] != deal["id"]])
        self.refresh_everything()

    def check_reminders(self):
        now = datetime.now()
        rows = self.store.deals()
        new_due = []
        changed = False
        for deal in rows:
            due = parse_jalali_datetime(deal.get("reminder_at", ""))
            if due is None:
                continue
            if due <= now and due >= self.app_started_at and deal.get("last_notified_at", "") != deal.get("reminder_at", ""):
                deal["last_notified_at"] = deal.get("reminder_at", "")
                changed = True
                new_due.append(deal)
        if new_due:
            if changed:
                self.store.save_deals(rows)
            for deal in new_due:
                self.raise_notification(deal)
            self.refresh_everything()
        self.check_birthdays()
        self.after(self.reminder_poll_ms, self.check_reminders)

    def raise_notification(self, deal: dict):
        title = f"یادآور معامله: {deal['title']}"
        message = f"مشتری: {deal['customer_name']} | وضعیت: {deal['status']}"
        launch = f"{APP_NOTIFY_ID}:deal|{self.profile.profile_id}|{deal['id']}"
        self.dispatch_notification(title, message, launch, lambda: self.open_notification_target("deal", deal["id"]))

    def check_birthdays(self):
        now = datetime.now()
        jalali_today = jdatetime.datetime.fromgregorian(datetime=now)
        if now.hour < 10:
            return
        rows = self.store.customers()
        changed = False
        for customer in rows:
            birthdate = normalize_jalali_date(customer.get("birthdate", ""))
            if not birthdate:
                continue
            try:
                birth = jdatetime.datetime.strptime(birthdate, "%Y/%m/%d")
            except ValueError:
                continue
            today_key = f"{jalali_today.year:04d}/{jalali_today.month:02d}/{jalali_today.day:02d}"
            if birth.month == jalali_today.month and birth.day == jalali_today.day and customer.get("birthday_notified_for", "") != today_key:
                customer["birthday_notified_for"] = today_key
                changed = True
                self.raise_birthday_notification(customer)
        if changed:
            self.store.save_customers(rows)

    def raise_birthday_notification(self, customer: dict):
        title = "یادآور تولد مشتری"
        message = f"{customer['name']} امروز تولد دارد"
        launch = f"{APP_NOTIFY_ID}:birthday|{self.profile.profile_id}|{customer['id']}"
        self.dispatch_notification(title, message, launch, lambda: self.open_notification_target("birthday", customer["id"]))

    def show_windows_notification(self, title: str, message: str, launch: str):
        try:
            toast = Notification(APP_NOTIFY_ID, title=title, msg=message, duration="short", launch=launch)
            toast.set_audio(audio.Default, loop=False)
            toast.show()
        except Exception:
            pass
        if winsound:
            try:
                winsound.Beep(1200, 250)
            except RuntimeError:
                pass

    def open_notification_target(self, target: str, record_id: str):
        if self.tray_icon:
            self.restore_from_tray()
        else:
            try:
                self.deiconify()
            except Exception:
                pass
        self.lift()
        self.focus_force()
        if target == "deal":
            self.mark_deal_notification_seen(record_id)
            self.open_deal_by_id(record_id)
        elif target == "birthday":
            self.mark_birthday_seen(record_id)
            self.open_customer_by_id(record_id)
        self.refresh_notifications()

    def dispatch_notification(self, title: str, message: str, launch: str, callback: Callable[[], None]):
        self.show_in_app_notification(title, message, callback)


def main():
    migrate_legacy_data_if_needed()
    DATA_ROOT.mkdir(parents=True, exist_ok=True)
    selector_root = tb.Window(themename="flatly")
    selector_root.title(APP_TITLE)
    selector_root.geometry("1x1+0+0")
    selector_root.withdraw()
    try:
        selector_root.attributes("-alpha", 0.0)
    except Exception:
        pass
    profile_store = ProfileStore()
    launch_arg = " ".join(sys.argv[1:]).strip().strip('"') if len(sys.argv) > 1 else ""
    launch_profile = None
    launch_target = None
    if launch_arg.startswith(f"{APP_NOTIFY_ID}:"):
        payload = launch_arg.split(":", 1)[1]
        parts = payload.split("|")
        if len(parts) == 3:
            launch_target, profile_id, record_id = parts
            launch_profile = profile_store.resolve_profile(profile_id)
            if launch_profile:
                app = CRMApp(selector_root, launch_profile)
                app.focus_force()
                app.lift()
                if launch_target == "deal":
                    app.mark_deal_notification_seen(record_id)
                    app.after(300, lambda: app.open_deal_by_id(record_id))
                elif launch_target == "birthday":
                    app.mark_birthday_seen(record_id)
                    app.after(300, lambda: app.open_customer_by_id(record_id))
                selector_root.mainloop()
                return
    dialog = ProfileDialog(selector_root, profile_store)
    if not dialog.result:
        selector_root.destroy()
        return
    app = CRMApp(selector_root, dialog.result)
    app.focus_force()
    app.lift()
    selector_root.mainloop()


if __name__ == "__main__":
    try:
        main()
    except KeyboardInterrupt:
        sys.exit(0)
